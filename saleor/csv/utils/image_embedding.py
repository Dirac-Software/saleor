"""Utilities for embedding images in Excel files during export."""

import logging
import os
from io import BytesIO
from typing import Any

from django.conf import settings
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter
from PIL import Image as PILImage

from ...core.http_client import HTTPClient

logger = logging.getLogger(__name__)


def convert_image_to_png(image_data: Any) -> BytesIO:
    """Convert any image format to PNG (which openpyxl supports).

    Args:
        image_data: BytesIO object or file path

    Returns:
        BytesIO object containing PNG image

    """
    # Open image with PIL
    if isinstance(image_data, str | os.PathLike):
        img = PILImage.open(image_data)
    else:
        img = PILImage.open(image_data)

    # Convert to RGB if necessary (for transparency handling)
    if img.mode in ("RGBA", "LA", "P"):
        # Create white background
        background = PILImage.new("RGB", img.size, (255, 255, 255))
        if img.mode == "P":
            img = img.convert("RGBA")  # type: ignore[assignment]
        background.paste(
            img,
            mask=img.split()[-1] if img.mode in ("RGBA", "LA") else None,
        )
        final_img = background
    elif img.mode != "RGB":
        final_img = img.convert("RGB")
    else:
        final_img = img

    # Save as PNG to BytesIO
    output = BytesIO()
    final_img.save(output, format="PNG")
    output.seek(0)
    return output


def download_and_convert_image(image_url: str, timeout: int = 15) -> BytesIO | None:
    """Download an image from a URL and convert it to PNG format.

    Args:
        image_url: URL of the image to download
        timeout: Request timeout in seconds

    Returns:
        BytesIO object containing PNG image, or None if download fails

    """
    try:
        response = HTTPClient.send_request("GET", image_url, timeout=timeout)
        response.raise_for_status()
        img_data = BytesIO(response.content)
        return convert_image_to_png(img_data)
    except Exception as e:
        logger.warning("Failed to download or convert image from %s: %s", image_url, e)
        return None


def load_local_image(image_path: str) -> BytesIO | None:
    """Load a local image file and convert it to PNG format.

    Args:
        image_path: Local file path to the image

    Returns:
        BytesIO object containing PNG image, or None if load fails

    """
    try:
        if os.path.exists(image_path):
            return convert_image_to_png(image_path)
        logger.warning("Local image file not found: %s", image_path)
        return None
    except Exception as e:
        logger.warning("Failed to load local image %s: %s", image_path, e)
        return None


def embed_images_in_excel(
    excel_path: str,
    image_columns: list[str],
    row_height: int = 80,
    col_width: int = 15,
):
    """Embed images into an Excel file for specified columns.

    Reads the Excel file, finds columns with image URLs, downloads the images,
    and embeds them directly in the cells.

    Args:
        excel_path: Path to the Excel file
        image_columns: List of column names that contain image URLs
        row_height: Height of rows with images in points
        col_width: Width of image columns in character units

    """
    logger.info("Starting image embedding for columns: %s", image_columns)

    # Load workbook
    wb = load_workbook(excel_path)
    ws = wb.active

    # Get headers from first row
    headers = []
    for cell in ws[1]:
        headers.append(cell.value)

    logger.info("Excel headers found: %s", headers)

    # Process each image column
    for image_column in image_columns:
        if image_column not in headers:
            logger.warning(
                "Image column '%s' not found in headers: %s", image_column, headers
            )
            continue

        # Find the column index (1-indexed for Excel)
        image_col_idx = headers.index(image_column) + 1
        image_col_letter = get_column_letter(image_col_idx)

        logger.info(
            "Processing image column '%s' at index %s", image_column, image_col_letter
        )

        # Set column width
        ws.column_dimensions[image_col_letter].width = col_width

        images_embedded = 0
        images_failed = 0

        # Process each data row (starting from row 2, after header)
        rows_processed = 0
        for row_num in range(2, ws.max_row + 1):
            cell = ws.cell(row=row_num, column=image_col_idx)
            image_value = cell.value

            # Log first few cell values for debugging
            if rows_processed < 3:
                logger.info("Row %s, cell value: %s", row_num, repr(image_value))

            rows_processed += 1

            # Skip if no image value
            if not image_value or not str(image_value).strip():
                continue

            # Handle multiple image URLs (separated by commas or newlines)
            image_urls_str = str(image_value).strip()
            # Split by comma, newline, or semicolon
            image_urls = [
                url.strip()
                for url in image_urls_str.replace("\n", ",")
                .replace(";", ",")
                .split(",")
                if url.strip()
            ]

            # Use the first image URL if multiple are present
            if image_urls:
                image_url = image_urls[0]

                try:
                    # Set row height
                    ws.row_dimensions[row_num].height = row_height

                    # Download or load image
                    png_data = None
                    if image_url.startswith(("http://", "https://")):
                        png_data = download_and_convert_image(image_url)
                    else:
                        # Try as local file path
                        full_path = os.path.join(settings.MEDIA_ROOT, image_url)
                        png_data = load_local_image(full_path)

                    if png_data:
                        # Clear the cell content (remove URL text)
                        cell.value = ""

                        # Create Excel image object
                        img = XLImage(png_data)

                        # Resize image to fit nicely in cell
                        max_height_pixels = (row_height * 1.33) - 4  # Small margin
                        max_width_pixels = (col_width * 7) - 4  # Small margin

                        # Scale to fit within both dimensions
                        height_scale = (
                            max_height_pixels / img.height
                            if img.height > max_height_pixels
                            else 1
                        )
                        width_scale = (
                            max_width_pixels / img.width
                            if img.width > max_width_pixels
                            else 1
                        )
                        scale = min(height_scale, width_scale, 1)  # Don't upscale

                        img.width = int(img.width * scale)
                        img.height = int(img.height * scale)

                        # Position image in cell
                        cell_address = f"{image_col_letter}{row_num}"
                        img.anchor = cell_address

                        # Add image to worksheet
                        ws.add_image(img)
                        images_embedded += 1
                    else:
                        # Error already logged in download/load function
                        images_failed += 1

                except Exception as e:
                    # Log error but continue processing
                    logger.warning(
                        "Row %s: Error embedding image from %s: %s",
                        row_num,
                        image_url,
                        e,
                    )
                    images_failed += 1

        logger.info(
            "Column '%s': Embedded %s images, %s failed",
            image_column,
            images_embedded,
            images_failed,
        )

    # Save workbook
    wb.save(excel_path)
