import datetime
import logging
import uuid
from tempfile import NamedTemporaryFile
from typing import IO, TYPE_CHECKING, Any, cast

import petl as etl
from django.conf import settings
from django.utils import timezone
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from ...core.db.connection import allow_writer
from ...core.utils.batches import queryset_in_batches
from ...discount.models import VoucherCode
from ...giftcard.models import GiftCard
from ...product.models import Product
from .. import FileTypes
from ..notifications import send_export_download_link_notification
from .image_embedding import embed_images_in_excel
from .product_headers import get_product_export_fields_and_headers_info
from .products_data import get_products_data
from .variant_compression import get_products_data_compressed

logger = logging.getLogger(__name__)


def convert_ids_to_proper_type(ids: list[str]) -> list[uuid.UUID | int]:
    """Convert string IDs to UUID or int based on format.

    Args:
        ids: List of string IDs

    Returns:
        List of converted IDs (UUID or int)

    """
    converted: list[uuid.UUID | int] = []
    for id_value in ids:
        if isinstance(id_value, str):
            # Try to convert to UUID, fall back to int
            try:
                converted.append(uuid.UUID(id_value))
            except ValueError:
                converted.append(int(id_value))
        else:
            converted.append(id_value)
    return converted


if TYPE_CHECKING:
    from django.db.models import QuerySet

    from ..models import ExportFile


BATCH_SIZE = 1000


def export_products(
    export_file: "ExportFile",
    scope: dict[str, str | dict],
    export_info: dict[str, list],
    file_type: str,
    delimiter: str = ",",
):
    from ...graphql.product.filters.product import ProductFilter

    file_name = get_filename("product", file_type)
    queryset = get_queryset(Product, ProductFilter, scope)

    (
        export_fields,
        file_headers,
        data_headers,
    ) = get_product_export_fields_and_headers_info(export_info)

    temporary_file = create_file_with_headers(file_headers, delimiter, file_type)

    # Check if we should compress variants
    compress_variants = export_info.get("compress_variants", False)

    if compress_variants:
        # Export with compressed variants (one row per product)
        export_products_compressed(
            queryset,
            export_info,
            data_headers,
            delimiter,
            temporary_file,
            file_type,
        )
    else:
        # Export with expanded variants (one row per variant)
        export_products_in_batches(
            queryset,
            export_info,
            set(export_fields),
            data_headers,
            delimiter,
            temporary_file,
            file_type,
        )

    # Embed images in Excel file if requested
    embed_images = export_info.get("embed_images", False)
    logger.info(
        "Image embedding: embed_images=%s, file_type=%s, file_headers=%s",
        embed_images,
        file_type,
        file_headers,
    )

    if embed_images and file_type == FileTypes.XLSX:
        # Identify which columns contain images
        image_columns = []
        if "product media" in file_headers:
            image_columns.append("product media")
        # Only include variant media if NOT compressing variants
        # (variant media doesn't make sense when there's one row per product)
        if "variant media" in file_headers and not compress_variants:
            image_columns.append("variant media")

        logger.info("Image columns to embed: %s", image_columns)

        # Embed images if there are any image columns
        if image_columns:
            logger.info("Starting image embedding for %s", temporary_file.name)
            embed_images_in_excel(temporary_file.name, image_columns)
            logger.info("Image embedding completed")
        else:
            logger.warning("No image columns found to embed")

    # Format currency columns in Excel file
    if file_type == FileTypes.XLSX and export_info.get("channels"):
        logger.info("Applying currency formatting to price columns")
        format_currency_columns(temporary_file.name, export_info)

    # Apply price list formatting if requested
    if file_type == FileTypes.XLSX and export_info.get("price_list_format"):
        logger.info("Applying price list formatting")
        format_as_price_list(temporary_file.name, export_info)

    save_csv_file_in_export_file(export_file, temporary_file, file_name)
    temporary_file.close()
    send_export_download_link_notification(export_file, "products")


def export_gift_cards(
    export_file: "ExportFile",
    scope: dict[str, str | dict],
    file_type: str,
    delimiter: str = ",",
):
    from ...graphql.giftcard.filters import GiftCardFilter

    file_name = get_filename("gift_card", file_type)

    queryset = get_queryset(GiftCard, GiftCardFilter, scope)
    # only unused gift cards codes can be exported
    queryset = queryset.filter(used_by_email__isnull=True)

    export_fields = ["code"]
    temporary_file = create_file_with_headers(export_fields, delimiter, file_type)

    export_gift_cards_in_batches(
        queryset,
        export_fields,
        delimiter,
        temporary_file,
        file_type,
    )

    save_csv_file_in_export_file(export_file, temporary_file, file_name)
    temporary_file.close()
    send_export_download_link_notification(export_file, "gift cards")


def export_voucher_codes(
    export_file: "ExportFile",
    file_type: str,
    voucher_id: int | None = None,
    ids: list[int] | None = None,
    delimiter: str = ",",
):
    file_name = get_filename("voucher_code", file_type)

    qs = VoucherCode.objects.using(settings.DATABASE_CONNECTION_REPLICA_NAME).all()
    if voucher_id:
        qs = VoucherCode.objects.using(
            settings.DATABASE_CONNECTION_REPLICA_NAME
        ).filter(voucher_id=voucher_id)
    if ids:
        qs = VoucherCode.objects.using(
            settings.DATABASE_CONNECTION_REPLICA_NAME
        ).filter(id__in=ids)

    export_fields = ["code"]
    temporary_file = create_file_with_headers(export_fields, delimiter, file_type)

    export_voucher_codes_in_batches(
        qs,
        export_fields,
        delimiter,
        temporary_file,
        file_type,
    )

    save_csv_file_in_export_file(export_file, temporary_file, file_name)
    temporary_file.close()
    send_export_download_link_notification(export_file, "voucher codes")


def get_filename(model_name: str, file_type: str) -> str:
    hash = uuid.uuid4()
    return "{}_data_{}_{}.{}".format(
        model_name, timezone.now().strftime("%d_%m_%Y_%H_%M_%S"), hash, file_type
    )


def get_queryset(model, filter, scope: dict[str, str | dict]) -> "QuerySet":
    queryset = model.objects.using(settings.DATABASE_CONNECTION_REPLICA_NAME).all()
    if "ids" in scope:
        queryset = model.objects.using(
            settings.DATABASE_CONNECTION_REPLICA_NAME
        ).filter(pk__in=scope["ids"])
    elif "filter" in scope:
        queryset = filter(data=parse_input(scope["filter"]), queryset=queryset).qs

    queryset = queryset.order_by("pk")

    return queryset


def parse_input(data: Any) -> dict[str, str | dict]:
    """Parse input into correct data types.

    Scope coming from Celery will be passed as strings.
    """
    if "attributes" in data:
        serialized_attributes = []

        for attr in data.get("attributes") or []:
            if "date_time" in attr:
                if gte := attr["date_time"].get("gte"):
                    attr["date_time"]["gte"] = datetime.datetime.fromisoformat(gte)
                if lte := attr["date_time"].get("lte"):
                    attr["date_time"]["lte"] = datetime.datetime.fromisoformat(lte)

            if "date" in attr:
                if gte := attr["date"].get("gte"):
                    attr["date"]["gte"] = datetime.date.fromisoformat(gte)
                if lte := attr["date"].get("lte"):
                    attr["date"]["lte"] = datetime.date.fromisoformat(lte)

            serialized_attributes.append(attr)

        if serialized_attributes:
            data["attributes"] = serialized_attributes

    return data


def format_as_price_list(excel_path: str, export_info: dict[str, list]) -> None:
    """Post-process Excel export into price list format.

    - Renames columns to clean names
    - Removes technical columns (IDs, internal fields)
    - Reorders columns to standard layout
    - Formats RRP with currency symbol

    Args:
        excel_path: Path to the Excel file
        export_info: Export info containing channel IDs and attributes

    """
    from ...channel.models import Channel

    wb = load_workbook(excel_path)
    ws = wb.active

    # Get current headers
    headers = [cell.value for cell in ws[1]]

    # Column renaming map
    RENAME_MAP = {
        "product media": "Image",
        "name": "Description",
        "category": "Category",
        "variants__size_quantity": "Sizes",
        "variants__total_quantity": "Qty",
    }

    # Add attribute renames (look for patterns)
    for _i, header in enumerate(headers):
        if header:
            header_str = str(header)
            header_lower = header_str.lower()
            # Rename attribute columns (case-insensitive, handle both hyphens and spaces)
            if (
                "product code" in header_lower or "product-code" in header_lower
            ) and "(product attribute)" in header_lower:
                RENAME_MAP[header] = "Product Code"
            elif "brand" in header_lower and "(product attribute)" in header_lower:
                RENAME_MAP[header] = "Brand"
            elif "rrp" in header_lower and "(product attribute)" in header_lower:
                RENAME_MAP[header] = "RRP"

    # Get channel info for price column renaming
    channel_ids = export_info.get("channels")
    channel_slug = None
    currency_code = None

    if channel_ids:
        converted_channel_ids = convert_ids_to_proper_type(channel_ids)
        channels = Channel.objects.using(
            settings.DATABASE_CONNECTION_REPLICA_NAME
        ).filter(pk__in=converted_channel_ids)
        if channels.exists():
            channel = channels.first()
            if channel:
                channel_slug = channel.slug
                currency_code = channel.currency_code

    # Rename price columns (remove channel prefix)
    if channel_slug:
        for _i, header in enumerate(headers):
            if header:
                header_str = str(header)
                # Rename "{channel} (channel price amount)" to "Price"
                if f"{channel_slug} (channel price amount)" == header_str:
                    RENAME_MAP[header] = "Price"

    # Columns to remove (by substring match)
    REMOVE_PATTERNS = [
        "id",  # Product/variant IDs
        "product type",
        "(channel variant currency code)",
        "(channel variant cost price)",
        "(channel published)",
        "(channel publication date)",  # Added
        "(channel published at)",  # Added
        "(channel searchable)",
        "(channel available for purchase)",
        "(channel product currency code)",
        "(channel variant preorder quantity threshold)",
    ]

    # Standard column order (columns that exist will be reordered to this)
    COLUMN_ORDER = [
        "Image",
        "Product Code",
        "Description",
        "Category",
        "Sizes",
        "Brand",
        "Qty",
        "RRP",
        "Price",
    ]

    # Step 1: Rename columns
    for cell in ws[1]:
        if cell.value in RENAME_MAP:
            cell.value = RENAME_MAP[cell.value]

    # Step 2: Remove unwanted columns
    headers_after_rename = [cell.value for cell in ws[1]]
    columns_to_delete = []

    for idx, header in enumerate(headers_after_rename, start=1):
        if header:
            header_str = str(header)
            # Check if header matches any remove pattern
            if any(pattern in header_str for pattern in REMOVE_PATTERNS):
                columns_to_delete.append(idx)

    # Delete columns in reverse order to maintain indices
    for col_idx in sorted(columns_to_delete, reverse=True):
        ws.delete_cols(col_idx, 1)

    # Step 3: Reorder columns to standard layout
    final_headers = [cell.value for cell in ws[1]]
    current_order = {header: idx for idx, header in enumerate(final_headers, start=1)}

    # Build new column order based on what exists
    new_order = []
    for col_name in COLUMN_ORDER:
        if col_name in current_order:
            new_order.append((col_name, current_order[col_name]))

    # Add any remaining columns not in COLUMN_ORDER (at the end)
    for header, idx in current_order.items():
        if header not in COLUMN_ORDER:
            new_order.append((header, idx))

    # Reorder by moving columns
    # This is complex in openpyxl, so we'll create a new sheet with correct order

    new_ws = wb.create_sheet("PriceList")

    # Copy headers in new order
    for new_col, (col_name, old_col) in enumerate(new_order, start=1):
        new_ws.cell(row=1, column=new_col).value = col_name

        # Copy all data for this column
        for row_num in range(2, ws.max_row + 1):
            old_cell = ws.cell(row=row_num, column=old_col)
            new_cell = new_ws.cell(row=row_num, column=new_col)
            new_cell.value = old_cell.value
            # Copy number format if it exists
            if old_cell.number_format:
                new_cell.number_format = old_cell.number_format

    # Copy row dimensions (heights) from old sheet to new sheet
    for row_num, row_dim in ws.row_dimensions.items():
        if row_dim.height:
            new_ws.row_dimensions[row_num].height = row_dim.height

    # Copy column dimensions (widths) - need to map old columns to new columns
    for new_col, (_col_name, old_col) in enumerate(new_order, start=1):
        old_col_letter = get_column_letter(old_col)
        new_col_letter = get_column_letter(new_col)
        if old_col_letter in ws.column_dimensions:
            old_col_dim = ws.column_dimensions[old_col_letter]
            if old_col_dim.width:
                new_ws.column_dimensions[new_col_letter].width = old_col_dim.width

    # Copy images from old sheet to new sheet
    # Images need to be repositioned to match the new column order
    if hasattr(ws, "_images") and ws._images:
        logger.info("Copying %s images to reformatted sheet", len(ws._images))
        from io import BytesIO

        from openpyxl.drawing.image import Image as XLImage

        for img in ws._images:
            # Get the anchor (cell reference) of the image
            anchor = img.anchor
            if hasattr(anchor, "_from"):
                # Get the column index from the image anchor
                old_col_idx = anchor._from.col + 1  # openpyxl uses 0-indexed columns
                row_idx = anchor._from.row + 1  # openpyxl uses 0-indexed rows

                # Find the new column index for this old column
                new_col_idx = None
                for new_col, (_col_name, old_col) in enumerate(new_order, start=1):
                    if old_col == old_col_idx:
                        new_col_idx = new_col
                        break

                # If we found a matching column, copy the image to the new position
                if new_col_idx:
                    new_col_letter = get_column_letter(new_col_idx)
                    new_anchor = f"{new_col_letter}{row_idx}"

                    # Create a new image instance with the same data
                    # Use the image's _data() method to get the raw image bytes
                    try:
                        img_bytes = BytesIO(img._data())
                        new_img = XLImage(img_bytes)
                        new_img.width = img.width
                        new_img.height = img.height
                        new_img.anchor = new_anchor

                        # Add to new worksheet
                        new_ws.add_image(new_img)
                        logger.debug(
                            "Copied image from %s%s to %s",
                            get_column_letter(old_col_idx),
                            row_idx,
                            new_anchor,
                        )
                    except Exception as e:
                        logger.warning(
                            "Failed to copy image from %s%s: %s",
                            get_column_letter(old_col_idx),
                            row_idx,
                            e,
                        )

    # Delete old sheet and rename new one
    wb.remove(ws)
    new_ws.title = "Sheet"

    # Step 4: Format RRP column with currency (if exists)
    headers_final = [cell.value for cell in new_ws[1]]
    logger.info(
        "RRP formatting check: currency_code=%s, headers=%s",
        currency_code,
        headers_final,
    )

    if currency_code and "RRP" in headers_final:
        # Currency formats
        CURRENCY_FORMATS = {
            "GBP": "[$£-809]#,##0.00",
            "USD": "[$$-409]#,##0.00",
            "EUR": "[$€-407]#,##0.00",
            "JPY": "[$¥-411]#,##0",
            "CNY": "[$¥-804]#,##0.00",
            "INR": "[$₹-439]#,##0.00",
            "AUD": "[$A$-C09]#,##0.00",
            "CAD": "[$C$-1009]#,##0.00",
        }

        number_format = CURRENCY_FORMATS.get(currency_code, "#,##0.00")
        logger.info("Applying currency format %s to RRP column", number_format)

        # Find RRP column
        rrp_col = headers_final.index("RRP") + 1
        formatted_count = 0
        for row in range(2, new_ws.max_row + 1):
            cell = new_ws.cell(row=row, column=rrp_col)
            if cell.value is not None:
                # Convert to float if it's a string
                try:
                    if isinstance(cell.value, str):
                        cell.value = float(cell.value)
                    cell.number_format = number_format
                    formatted_count += 1
                except (ValueError, TypeError):
                    # Keep original value if conversion fails
                    logger.warning(
                        "Could not convert RRP value to number: %s", cell.value
                    )

        logger.info("Formatted %s RRP cells with currency", formatted_count)

    wb.save(excel_path)
    logger.info("Applied price list formatting to export")


def format_currency_columns(excel_path: str, export_info: dict[str, list]) -> None:
    """Format price columns in Excel with currency symbols based on channel currency.

    Args:
        excel_path: Path to the Excel file
        export_info: Export info containing channel IDs

    """
    from ...channel.models import Channel

    channel_ids = export_info.get("channels")
    if not channel_ids:
        return

    # Convert channel IDs to proper type
    converted_channel_ids = convert_ids_to_proper_type(channel_ids)

    # Get channel currencies
    channels = Channel.objects.using(settings.DATABASE_CONNECTION_REPLICA_NAME).filter(
        pk__in=converted_channel_ids
    )

    # Currency code to Excel number format mapping
    CURRENCY_FORMATS = {
        "GBP": "[$£-809]#,##0.00",
        "USD": "[$$-409]#,##0.00",
        "EUR": "[$€-407]#,##0.00",
        "JPY": "[$¥-411]#,##0",
        "CNY": "[$¥-804]#,##0.00",
        "INR": "[$₹-439]#,##0.00",
        "AUD": "[$A$-C09]#,##0.00",
        "CAD": "[$C$-1009]#,##0.00",
        "CHF": "[$CHF-807]#,##0.00",
        "SEK": "[$kr-41D]#,##0.00",
        "NOK": "[$kr-414]#,##0.00",
        "DKK": "[$kr-406]#,##0.00",
        "PLN": "[$zł-415]#,##0.00",
    }

    # Load workbook
    wb = load_workbook(excel_path)
    ws = wb.active

    # Get headers
    headers = [cell.value for cell in ws[1]]

    # Track total columns formatted
    total_columns_formatted = 0

    # Find price-related columns for each channel
    for channel in channels:
        currency_code = channel.currency_code
        channel_slug = channel.slug
        number_format = CURRENCY_FORMATS.get(currency_code, "#,##0.00")

        # Find columns that contain price amounts for this channel
        price_columns = []
        for idx, header in enumerate(headers):
            if header and channel_slug in str(header):
                # Match patterns like "default (channel price amount)"
                if "price amount" in str(header).lower():
                    price_columns.append(idx + 1)  # 1-indexed
                # Also format cost price
                elif "cost price" in str(header).lower():
                    price_columns.append(idx + 1)

        # Apply formatting to these columns
        for col_idx in price_columns:
            for row in range(2, ws.max_row + 1):  # Skip header
                cell = ws.cell(row=row, column=col_idx)
                if cell.value is not None:
                    cell.number_format = number_format

        total_columns_formatted += len(price_columns)

    # Save workbook
    wb.save(excel_path)
    logger.info(
        "Applied currency formatting to %s price columns", total_columns_formatted
    )


def create_file_with_headers(file_headers: list[str], delimiter: str, file_type: str):
    table = etl.wrap([file_headers])

    if file_type == FileTypes.CSV:
        temp_file = NamedTemporaryFile("ab+", suffix=".csv")
        etl.tocsv(table, temp_file.name, delimiter=delimiter)
    else:
        temp_file = NamedTemporaryFile("ab+", suffix=".xlsx")
        etl.io.xlsx.toxlsx(table, temp_file.name)

    return temp_file


def export_products_in_batches(
    queryset: "QuerySet",
    export_info: dict[str, list],
    export_fields: set[str],
    headers: list[str],
    delimiter: str,
    temporary_file: Any,
    file_type: str,
):
    warehouses = export_info.get("warehouses")
    attributes = export_info.get("attributes")
    channels = export_info.get("channels")

    for batch_pks in queryset_in_batches(queryset, BATCH_SIZE):
        product_batch = (
            Product.objects.using(settings.DATABASE_CONNECTION_REPLICA_NAME)
            .filter(pk__in=batch_pks)
            .prefetch_related(
                "attributevalues",
                "variants",
                "collections",
                "media",
                "product_type",
                "category",
            )
        )
        export_data = get_products_data(
            product_batch, export_fields, attributes, warehouses, channels
        )

        append_to_file(
            cast(list[dict[str, str | bool | float | int | None]], export_data),
            headers,
            temporary_file,
            file_type,
            delimiter,
        )


def export_products_compressed(
    queryset: "QuerySet",
    export_info: dict[str, list],
    headers: list[str],
    delimiter: str,
    temporary_file: Any,
    file_type: str,
):
    """Export products with compressed variants (one row per product)."""
    from collections import ChainMap

    from . import ProductExportFields

    warehouses = export_info.get("warehouses")
    attributes = export_info.get("attributes")
    channels = export_info.get("channels")
    requested_fields = export_info.get("fields", [])

    # Build export_fields set from requested fields
    fields_mapping = dict(
        ChainMap(*reversed(ProductExportFields.HEADERS_TO_FIELDS_MAPPING.values()))
    )
    export_fields_set = set()
    if requested_fields:
        for field in requested_fields:
            if field in fields_mapping:
                export_fields_set.add(fields_mapping[field])

    for batch_pks in queryset_in_batches(queryset, BATCH_SIZE):
        product_batch = (
            Product.objects.using(settings.DATABASE_CONNECTION_REPLICA_NAME)
            .filter(pk__in=batch_pks)
            .prefetch_related(
                "attributevalues",
                "variants",
                "collections",
                "media",
                "product_type",
                "category",
            )
        )

        # Use compressed data function
        export_data = get_products_data_compressed(
            product_batch,
            export_fields_set,  # Pass proper export fields
            attributes,
            warehouses,
            channels,
            requested_fields,
        )

        append_to_file(export_data, headers, temporary_file, file_type, delimiter)


def export_gift_cards_in_batches(
    queryset: "QuerySet",
    export_fields: list[str],
    delimiter: str,
    temporary_file: Any,
    file_type: str,
):
    for batch_pks in queryset_in_batches(queryset, BATCH_SIZE):
        gift_card_batch = GiftCard.objects.using(
            settings.DATABASE_CONNECTION_REPLICA_NAME
        ).filter(pk__in=batch_pks)

        export_data = list(gift_card_batch.values(*export_fields))

        append_to_file(export_data, export_fields, temporary_file, file_type, delimiter)


def export_voucher_codes_in_batches(
    queryset: "QuerySet",
    export_fields: list[str],
    delimiter: str,
    temporary_file: Any,
    file_type: str,
):
    for batch_pks in queryset_in_batches(queryset, BATCH_SIZE):
        voucher_codes_batch = VoucherCode.objects.using(
            settings.DATABASE_CONNECTION_REPLICA_NAME
        ).filter(pk__in=batch_pks)

        export_data = list(voucher_codes_batch.values(*export_fields))

        append_to_file(export_data, export_fields, temporary_file, file_type, delimiter)


def append_to_file(
    export_data: list[dict[str, str | bool | float | int | None]],
    headers: list[str],
    temporary_file: Any,
    file_type: str,
    delimiter: str,
):
    table = etl.fromdicts(export_data, header=headers, missing="")

    if file_type == FileTypes.CSV:
        etl.io.csv.appendcsv(table, temporary_file.name, delimiter=delimiter)
    else:
        etl.io.xlsx.appendxlsx(table, temporary_file.name)


@allow_writer()
def save_csv_file_in_export_file(
    export_file: "ExportFile", temporary_file: IO[bytes], file_name: str
):
    export_file.content_file.save(file_name, temporary_file)
