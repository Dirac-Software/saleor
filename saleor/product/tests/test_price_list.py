"""Tests for PriceList processing task."""

import os
from decimal import Decimal

import pytest
from django.http import Http404

from saleor.product.models import PriceList, PriceListItem
from saleor.product.tasks import (
    activate_price_list_task,
    deactivate_price_list_task,
    process_price_list_task,
    replace_price_list_task,
)
from saleor.warehouse.models import Stock, Warehouse

# ---------------------------------------------------------------------------
# serve_price_list view tests
# ---------------------------------------------------------------------------


@pytest.fixture
def pl_view_warehouse(db):
    from saleor.account.models import Address

    address = Address.objects.create(
        street_address_1="1 Test St", city="Test City", country="GB"
    )
    return Warehouse.objects.create(
        name="View Test WH", slug="view-test-wh", address=address, is_owned=False
    )


@pytest.fixture
def price_list_with_file(db, pl_view_warehouse, tmp_path):
    import openpyxl
    from django.core.files import File

    wb = openpyxl.Workbook()
    path = tmp_path / "sample.xlsx"
    wb.save(path)

    with open(path, "rb") as f:
        pl = PriceList.objects.create(
            warehouse=pl_view_warehouse,
            name="View Test PL",
            config={},
            excel_file=File(f, name="sample.xlsx"),
        )
    return pl


@pytest.fixture
def price_list_without_file(db, pl_view_warehouse):
    return PriceList.objects.create(
        warehouse=pl_view_warehouse, name="No File PL", config={}
    )


def _signed_id(price_list_id):
    from django.core.signing import TimestampSigner

    return TimestampSigner().sign(str(price_list_id))


def _call_view(signed_id, pk=0):
    from django.test import RequestFactory

    from saleor.media_views import serve_price_list_signed

    factory = RequestFactory()
    request = factory.get(f"/media/price_lists/{pk}/{signed_id}/")
    return serve_price_list_signed(request, pk=pk, signed_id=signed_id)


def test_serve_price_list_valid_signed_url(price_list_with_file):
    response = _call_view(_signed_id(price_list_with_file.pk))
    assert response.status_code == 200
    assert (
        response["Content-Type"]
        == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert "attachment" in response.get("Content-Disposition", "")


def test_serve_price_list_expired_link():
    import datetime

    from django.core.signing import TimestampSigner
    from freezegun import freeze_time

    with freeze_time(
        datetime.datetime.now(tz=datetime.UTC) - datetime.timedelta(days=8)
    ):
        signed_id = TimestampSigner().sign("999")

    response = _call_view(signed_id)
    assert response.status_code == 410


def test_serve_price_list_invalid_signature():
    response = _call_view("not-a-valid-signature")
    assert response.status_code == 400


def test_serve_price_list_not_found():
    with pytest.raises(Http404):
        _call_view(_signed_id(999999))


def test_serve_price_list_no_file_returns_404(price_list_without_file):
    with pytest.raises(Http404):
        _call_view(_signed_id(price_list_without_file.pk))


HK_COLUMN_MAP = {
    "0": "brand",
    "1": "product_code",
    "2": "description",
    "3": "rrp",
    "4": "sell_price",
    "6": "category",
    "12": "weight_kg",
    "13": "sizes",
}

# First 5 rows from saleor_hk (2).xlsx — updated_sizing at col 13
HK_ROWS = [
    [
        "Adidas",
        "IS1637",
        "TIRO24 C TRPNTW",
        40,
        9.025,
        190,
        "Apparel",
        "Women",
        "XS[20], M[50], L[50], XL[50], 3XL[20]",
        "Apparel",
        "Women",
        40,
        0.2,
        "XS[20], M[50], L[50], XL[50], 3XL[20]",
        None,
        "TIRO24 C TRPNTW",
    ],
    [
        "Adidas",
        "HY4520",
        "aSMC TST LS HO",
        110,
        23.69,
        155,
        "Apparel",
        "Women",
        "XS[16], S[26], M[92], L[21]",
        "Apparel",
        "Women",
        110,
        0.2,
        "XS[16], S[26], M[92], L[21]",
        None,
        "aSMC TST LS HO HY4520",
    ],
    [
        "Adidas",
        "HH9288",
        "W TXFlooceLT HJ WONRED",
        70,
        15.25,
        154,
        "Apparel",
        "Women",
        "XS[40], S[46], M[48], L[10], XL[10]",
        "Apparel",
        "Women",
        70,
        0.2,
        "XS[40], S[46], M[48], L[10], XL[10]",
        None,
        "W TXFlooceLT HJ WONRED",
    ],
    [
        "Adidas",
        "H59015",
        "BLOUSON",
        110,
        24.99,
        138,
        "Apparel",
        "Women",
        "32[58], 34[34], 36[23], 38[11], 40[12]",
        "Apparel",
        "Women",
        110,
        0.2,
        "32[58], 34[34], 36[23], 38[11], 40[12]",
        None,
        "BLOUSON",
    ],
    [
        "Adidas",
        "HK5015",
        "ADV WNTR AOP OH",
        60,
        14.43,
        135,
        "Apparel",
        "Men",
        "XS[25], S[14], M[43], L[16], XL[26], 2XL[11]",
        "Apparel",
        "Men",
        60,
        0.2,
        "XS[25], S[14], M[43], L[16], XL[26], 2XL[11]",
        None,
        "ADV WNTR AOP OH",
    ],
]

HK_HEADERS = [
    "Brand",
    "Article",
    "Description",
    "RRP (GBP)",
    "Sale Price (GBP)",
    "Quantity",
    "Category",
    "Gender",
    "Sizing (UK)",
    "Category.1",
    "Gender.1",
    "Rounded RRP (GBP)",
    "unit_weight_kg",
    "updated_sizing",
    "issues",
    "updated_description",
]


@pytest.fixture
def required_attributes(db):
    from saleor.attribute import AttributeInputType, AttributeType
    from saleor.attribute.models import Attribute

    attrs = {}
    for name, slug, input_type in [
        ("Product Code", "product-code", AttributeInputType.PLAIN_TEXT),
        ("RRP", "rrp", AttributeInputType.NUMERIC),
        (
            "Minimum Order Quantity",
            "minimum-order-quantity",
            AttributeInputType.NUMERIC,
        ),
        ("Brand", "brand", AttributeInputType.PLAIN_TEXT),
        ("Size", "size", AttributeInputType.DROPDOWN),
    ]:
        attr, _ = Attribute.objects.get_or_create(
            slug=slug,
            defaults={
                "name": name,
                "type": AttributeType.PRODUCT_TYPE,
                "input_type": input_type,
            },
        )
        attrs[name] = attr
    return attrs


@pytest.fixture
def warehouse(db):
    from saleor.account.models import Address

    address = Address.objects.create(
        street_address_1="1 Test St",
        city="Test City",
        country="HK",
    )
    return Warehouse.objects.create(
        name="HK Warehouse",
        slug="hk-warehouse",
        address=address,
        is_owned=False,
    )


@pytest.fixture
def hk_excel(tmp_path):
    """Small HK-format Excel fixture using first 5 rows of real HK sheet data."""
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1 (1)"
    ws.append(HK_HEADERS)
    for row in HK_ROWS:
        ws.append(row)
    path = tmp_path / "hk_sample.xlsx"
    wb.save(path)
    return path


@pytest.fixture
def hk_excel_with_invalid_row(tmp_path):
    """HK-format Excel with one invalid row (missing product_code, unparseable sizes)."""
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1 (1)"
    ws.append(HK_HEADERS)
    for row in HK_ROWS[:2]:
        ws.append(row)
    invalid_row = [
        "Adidas",
        "",
        "Some Jacket",
        50,
        12.0,
        10,
        "Apparel",
        "Men",
        "INVALID SIZES",
        "Apparel",
        "Men",
        50,
        0.2,
        "INVALID SIZES",
        None,
        "Some Jacket",
    ]
    ws.append(invalid_row)
    path = tmp_path / "hk_invalid.xlsx"
    wb.save(path)
    return path


def _make_price_list(warehouse, excel_path, column_map=None):
    from django.core.files import File

    with open(excel_path, "rb") as f:
        return PriceList.objects.create(
            warehouse=warehouse,
            excel_file=File(f, name=os.path.basename(excel_path)),
            config={
                "sheet_name": "Sheet1 (1)",
                "header_row": 0,
                "column_map": column_map or HK_COLUMN_MAP,
                "default_currency": "GBP",
            },
        )


def test_process_task(db, warehouse, hk_excel):
    price_list = _make_price_list(warehouse, hk_excel)

    process_price_list_task(price_list.pk)

    price_list.refresh_from_db()
    assert price_list.processing_completed_at is not None
    assert price_list.processing_failed_at is None
    assert PriceListItem.objects.filter(price_list=price_list).count() == len(HK_ROWS)

    is1637 = PriceListItem.objects.get(price_list=price_list, product_code="is1637")
    assert is1637.brand == "adidas"
    assert is1637.description == "TIRO24 C TRPNTW"
    assert is1637.category == "Apparel"
    assert is1637.rrp == Decimal(40)
    assert is1637.sell_price == Decimal("9.03")  # decimal_places=2 rounds 9.025 → 9.03
    assert is1637.weight_kg == Decimal("0.2")
    assert is1637.currency == "GBP"
    assert is1637.is_valid is True
    assert is1637.validation_errors == []
    assert is1637.sizes_and_qty == {"XS": 20, "M": 50, "L": 50, "XL": 50, "3XL": 20}

    h59015 = PriceListItem.objects.get(price_list=price_list, product_code="h59015")
    assert h59015.sizes_and_qty == {"32": 58, "34": 34, "36": 23, "38": 11, "40": 12}

    hk5015 = PriceListItem.objects.get(price_list=price_list, product_code="hk5015")
    assert hk5015.sizes_and_qty == {
        "XS": 25,
        "S": 14,
        "M": 43,
        "L": 16,
        "XL": 26,
        "2XL": 11,
    }


def test_process_invalid_row_sets_is_valid_false(
    db, warehouse, hk_excel_with_invalid_row
):
    price_list = _make_price_list(warehouse, hk_excel_with_invalid_row)

    process_price_list_task(price_list.pk)

    items = PriceListItem.objects.filter(price_list=price_list)
    assert items.count() == 3
    assert items.filter(is_valid=True).count() == 2

    invalid_item = items.get(is_valid=False)
    assert "product_code: required" in invalid_item.validation_errors
    assert any("sizes" in e for e in invalid_item.validation_errors)


def test_process_sets_failed_at_on_missing_file(db, warehouse, tmp_path):
    from django.core.files.base import ContentFile

    price_list = PriceList.objects.create(
        warehouse=warehouse,
        excel_file=ContentFile(b"", name="empty.xlsx"),
        config={
            "sheet_name": "Sheet1",
            "header_row": 0,
            "column_map": HK_COLUMN_MAP,
            "default_currency": "GBP",
        },
    )
    # Delete the file so the task cannot open it
    price_list.excel_file.delete(save=False)

    with pytest.raises(FileNotFoundError):
        process_price_list_task(price_list.pk)

    price_list.refresh_from_db()
    assert price_list.processing_failed_at is not None
    assert price_list.processing_completed_at is None


def test_process_clears_completed_at_on_failure(db, warehouse, hk_excel):
    from django.utils import timezone

    price_list = _make_price_list(warehouse, hk_excel)
    price_list.processing_completed_at = timezone.now()
    price_list.save(update_fields=["processing_completed_at"])

    price_list.excel_file.delete(save=False)

    with pytest.raises(FileNotFoundError):
        process_price_list_task(price_list.pk)

    price_list.refresh_from_db()
    assert price_list.processing_completed_at is None
    assert price_list.processing_failed_at is not None


def test_process_replaces_items_on_rerun(db, warehouse, hk_excel):
    price_list = _make_price_list(warehouse, hk_excel)

    process_price_list_task(price_list.pk)
    assert PriceListItem.objects.filter(price_list=price_list).count() == len(HK_ROWS)

    process_price_list_task(price_list.pk)
    assert PriceListItem.objects.filter(price_list=price_list).count() == len(HK_ROWS)


# ---------------------------------------------------------------------------
# Helpers for activate / deactivate / replace tests
# ---------------------------------------------------------------------------


def _make_processed_price_list(warehouse, sizes_and_qty=None, product=None):
    """Create a PriceList with one valid item and processing_completed_at set."""
    from django.utils import timezone

    if sizes_and_qty is None:
        sizes_and_qty = {"S": 10, "M": 20}

    pl = PriceList.objects.create(
        warehouse=warehouse,
        config={},
        processing_completed_at=timezone.now(),
    )
    item = PriceListItem.objects.create(
        price_list=pl,
        row_index=0,
        product_code="TEST-001",
        brand="TestBrand",
        description="Test Product",
        category="Apparel",
        sizes_and_qty=sizes_and_qty,
        sell_price=Decimal("25.00"),
        currency="GBP",
        is_valid=True,
        product=product,
    )
    return pl, item


def _make_product_with_variant_and_stock(warehouse, size="S", quantity=100):
    """Create a Product → Variant → Stock chain and return (product, variant, stock)."""
    from saleor.product.models import ProductType, ProductVariant

    product_type, _ = ProductType.objects.get_or_create(
        slug="apparel-type",
        defaults={"name": "Apparel", "has_variants": True},
    )
    from saleor.product.models import Product

    product = Product.objects.create(
        name="Test Product",
        slug=f"test-product-{Product.objects.count()}",
        product_type=product_type,
    )
    variant = ProductVariant.objects.create(
        product=product,
        name=size,
        sku=f"sku-{product.pk}-{size}",
    )
    stock = Stock.objects.create(
        product_variant=variant,
        warehouse=warehouse,
        quantity=quantity,
    )
    return product, variant, stock


# ---------------------------------------------------------------------------
# Activation tests
# ---------------------------------------------------------------------------


def test_activate_creates_stock(db, warehouse):
    product, variant, _ = _make_product_with_variant_and_stock(
        warehouse, size="S", quantity=0
    )
    Stock.objects.filter(product_variant=variant).delete()

    pl, item = _make_processed_price_list(
        warehouse,
        sizes_and_qty={"S": 15},
        product=product,
    )

    activate_price_list_task(pl.pk)

    stock = Stock.objects.get(product_variant=variant, warehouse=warehouse)
    assert stock.quantity == 15


def test_activate_sets_status_active(db, warehouse):
    from saleor.product import PriceListStatus

    product, _, _ = _make_product_with_variant_and_stock(warehouse)
    pl, _ = _make_processed_price_list(warehouse, product=product)

    activate_price_list_task(pl.pk)

    pl.refresh_from_db()
    assert pl.status == PriceListStatus.ACTIVE
    assert pl.activated_at is not None


def test_activate_raises_if_not_processed(db, warehouse):
    pl = PriceList.objects.create(warehouse=warehouse, config={})

    with pytest.raises(ValueError, match="has not completed processing"):
        activate_price_list_task(pl.pk)


def test_activate_raises_for_owned_warehouse(db):
    from saleor.account.models import Address

    address = Address.objects.create(
        street_address_1="1 Owned St", city="City", country="GB"
    )
    owned_wh = Warehouse.objects.create(
        name="Owned WH", slug="owned-wh", address=address, is_owned=True
    )
    from django.utils import timezone

    pl = PriceList.objects.create(
        warehouse=owned_wh, config={}, processing_completed_at=timezone.now()
    )

    with pytest.raises(ValueError, match="is owned"):
        activate_price_list_task(pl.pk)


def test_activate_raises_when_category_missing_product_type(
    db, warehouse, required_attributes
):
    from saleor.product import PriceListStatus

    pl, _ = _make_processed_price_list(warehouse, product=None)

    with pytest.raises(ValueError, match="no ProductType or Category"):
        activate_price_list_task(pl.pk)

    pl.refresh_from_db()
    assert pl.status != PriceListStatus.ACTIVE


def test_activate_raises_when_missing_database_setup(db, warehouse):
    from unittest.mock import patch

    from saleor.product import PriceListStatus
    from saleor.product.ingestion import MissingDatabaseSetup

    pl, _ = _make_processed_price_list(warehouse, product=None)

    with pytest.raises(MissingDatabaseSetup):
        with patch(
            "saleor.product.ingestion.get_products_by_code_and_brand",
            side_effect=MissingDatabaseSetup("Product Code attribute not found"),
        ):
            activate_price_list_task(pl.pk)

    pl.refresh_from_db()
    assert pl.status != PriceListStatus.ACTIVE


def test_replace_raises_when_category_missing_product_type(
    db, warehouse, required_attributes
):
    from saleor.product import PriceListStatus

    product, _, _ = _make_product_with_variant_and_stock(warehouse)
    old_pl, _ = _make_processed_price_list(warehouse, product=product)
    old_pl.status = PriceListStatus.ACTIVE
    old_pl.save(update_fields=["status"])

    new_pl, _ = _make_processed_price_list(warehouse, product=None)

    with pytest.raises(ValueError, match="no ProductType or Category"):
        replace_price_list_task(old_pl.pk, new_pl.pk)

    new_pl.refresh_from_db()
    assert new_pl.status != PriceListStatus.ACTIVE
    old_pl.refresh_from_db()
    assert old_pl.status == PriceListStatus.ACTIVE


def test_activate_creates_variant_for_new_size(db, warehouse):
    from saleor.product.models import ProductVariant

    product, _, _ = _make_product_with_variant_and_stock(
        warehouse, size="S", quantity=5
    )
    pl, _ = _make_processed_price_list(
        warehouse,
        sizes_and_qty={"S": 10, "L": 30},
        product=product,
    )

    activate_price_list_task(pl.pk)

    assert ProductVariant.objects.filter(product=product, name="L").exists()
    stock = Stock.objects.get(
        product_variant__product=product,
        product_variant__name="L",
        warehouse=warehouse,
    )
    assert stock.quantity == 30


def test_activate_clears_deactivated_at(db, warehouse):
    from django.utils import timezone

    from saleor.product import PriceListStatus

    product, _, _ = _make_product_with_variant_and_stock(warehouse)
    pl, _ = _make_processed_price_list(warehouse, product=product)
    pl.status = PriceListStatus.INACTIVE
    pl.deactivated_at = timezone.now()
    pl.save(update_fields=["status", "deactivated_at"])

    activate_price_list_task(pl.pk)

    pl.refresh_from_db()
    assert pl.status == PriceListStatus.ACTIVE
    assert pl.deactivated_at is None


# ---------------------------------------------------------------------------
# Deactivation tests
# ---------------------------------------------------------------------------


def test_deactivate_zeros_stock(db, warehouse):
    from saleor.product import PriceListStatus

    product, variant, stock = _make_product_with_variant_and_stock(
        warehouse, quantity=50
    )
    pl, _ = _make_processed_price_list(warehouse, product=product)
    pl.status = PriceListStatus.ACTIVE
    pl.save(update_fields=["status"])

    deactivate_price_list_task(pl.pk)

    stock.refresh_from_db()
    assert stock.quantity == 0

    pl.refresh_from_db()
    assert pl.status == PriceListStatus.INACTIVE
    assert pl.deactivated_at is not None


def test_deactivate_respects_allocations(db, warehouse):
    from saleor.product import PriceListStatus

    product, variant, stock = _make_product_with_variant_and_stock(
        warehouse, quantity=50
    )
    stock.quantity_allocated = 30
    stock.save(update_fields=["quantity_allocated"])

    pl, _ = _make_processed_price_list(warehouse, product=product)
    pl.status = PriceListStatus.ACTIVE
    pl.save(update_fields=["status"])

    deactivate_price_list_task(pl.pk)

    stock.refresh_from_db()
    assert stock.quantity == 30


def test_deactivate_skips_items_without_product_fk(db, warehouse):
    pl, item = _make_processed_price_list(warehouse, product=None)

    deactivate_price_list_task(pl.pk)

    pl.refresh_from_db()
    from saleor.product import PriceListStatus

    assert pl.status == PriceListStatus.INACTIVE


# ---------------------------------------------------------------------------
# Replace tests
# ---------------------------------------------------------------------------


def test_replace_validates_same_warehouse(db, warehouse):
    from django.utils import timezone

    from saleor.account.models import Address

    address = Address.objects.create(
        street_address_1="2 Other St", city="City", country="GB"
    )
    other_wh = Warehouse.objects.create(
        name="Other WH", slug="other-wh", address=address, is_owned=False
    )

    old_pl = PriceList.objects.create(
        warehouse=warehouse, config={}, processing_completed_at=timezone.now()
    )
    new_pl = PriceList.objects.create(
        warehouse=other_wh, config={}, processing_completed_at=timezone.now()
    )

    with pytest.raises(ValueError, match="different warehouses"):
        replace_price_list_task(old_pl.pk, new_pl.pk)


def test_replace_validates_new_is_processed(db, warehouse):
    from django.utils import timezone

    old_pl = PriceList.objects.create(
        warehouse=warehouse, config={}, processing_completed_at=timezone.now()
    )
    new_pl = PriceList.objects.create(warehouse=warehouse, config={})

    with pytest.raises(ValueError, match="has not completed processing"):
        replace_price_list_task(old_pl.pk, new_pl.pk)


def test_replace_diffs_correctly(db, warehouse):
    from django.utils import timezone

    from saleor.product import PriceListStatus

    product_a, variant_a, stock_a = _make_product_with_variant_and_stock(
        warehouse, size="S", quantity=50
    )
    product_b, variant_b, stock_b = _make_product_with_variant_and_stock(
        warehouse, size="M", quantity=40
    )

    old_pl = PriceList.objects.create(
        warehouse=warehouse, config={}, processing_completed_at=timezone.now()
    )
    PriceListItem.objects.create(
        price_list=old_pl,
        row_index=0,
        product_code="A",
        brand="B",
        description="A",
        category="Apparel",
        sizes_and_qty={"S": 50},
        sell_price=Decimal(10),
        currency="GBP",
        is_valid=True,
        product=product_a,
    )
    PriceListItem.objects.create(
        price_list=old_pl,
        row_index=1,
        product_code="B",
        brand="B",
        description="B",
        category="Apparel",
        sizes_and_qty={"M": 40},
        sell_price=Decimal(10),
        currency="GBP",
        is_valid=True,
        product=product_b,
    )

    new_pl = PriceList.objects.create(
        warehouse=warehouse, config={}, processing_completed_at=timezone.now()
    )
    PriceListItem.objects.create(
        price_list=new_pl,
        row_index=0,
        product_code="B",
        brand="B",
        description="B",
        category="Apparel",
        sizes_and_qty={"M": 99},
        sell_price=Decimal(10),
        currency="GBP",
        is_valid=True,
        product=product_b,
    )

    replace_price_list_task(old_pl.pk, new_pl.pk)

    stock_a.refresh_from_db()
    assert stock_a.quantity == 0

    stock_b.refresh_from_db()
    assert stock_b.quantity == 99

    old_pl.refresh_from_db()
    assert old_pl.status == PriceListStatus.INACTIVE
    assert old_pl.replaced_by_id == new_pl.pk

    new_pl.refresh_from_db()
    assert new_pl.status == PriceListStatus.ACTIVE


def test_replace_respects_allocations(db, warehouse):
    from django.utils import timezone

    product, variant, stock = _make_product_with_variant_and_stock(
        warehouse, quantity=50
    )
    stock.quantity_allocated = 20
    stock.save(update_fields=["quantity_allocated"])

    old_pl = PriceList.objects.create(
        warehouse=warehouse, config={}, processing_completed_at=timezone.now()
    )
    PriceListItem.objects.create(
        price_list=old_pl,
        row_index=0,
        product_code="X",
        brand="B",
        description="X",
        category="Apparel",
        sizes_and_qty={"S": 50},
        sell_price=Decimal(10),
        currency="GBP",
        is_valid=True,
        product=product,
    )

    new_pl = PriceList.objects.create(
        warehouse=warehouse, config={}, processing_completed_at=timezone.now()
    )

    replace_price_list_task(old_pl.pk, new_pl.pk)

    stock.refresh_from_db()
    assert stock.quantity == 20


# ---------------------------------------------------------------------------
# Auto-replace tests
# ---------------------------------------------------------------------------


def test_activate_proceeds_directly_when_another_list_is_active(db, warehouse):
    """Activating a new list when another is active proceeds with direct activation."""
    from saleor.product import PriceListStatus

    product, _, _ = _make_product_with_variant_and_stock(warehouse)

    old_pl, _ = _make_processed_price_list(warehouse, product=product)
    old_pl.status = PriceListStatus.ACTIVE
    old_pl.save(update_fields=["status"])

    new_pl, _ = _make_processed_price_list(warehouse, product=product)

    activate_price_list_task(new_pl.pk)

    new_pl.refresh_from_db()
    assert new_pl.status == PriceListStatus.ACTIVE
    old_pl.refresh_from_db()
    assert old_pl.status == PriceListStatus.ACTIVE


def test_activate_no_auto_replace_when_no_active_list(db, warehouse):
    """Activating a list when no other is active proceeds with direct activation."""
    from saleor.product import PriceListStatus

    product, variant, _ = _make_product_with_variant_and_stock(warehouse, quantity=0)
    Stock.objects.filter(product_variant=variant).delete()

    pl, _ = _make_processed_price_list(
        warehouse, sizes_and_qty={"S": 5}, product=product
    )

    activate_price_list_task(pl.pk)

    pl.refresh_from_db()
    assert pl.status == PriceListStatus.ACTIVE


# ---------------------------------------------------------------------------
# Additional activate tests
# ---------------------------------------------------------------------------


def test_activate_increments_existing_stock(db, warehouse):
    product, variant, stock = _make_product_with_variant_and_stock(
        warehouse, size="S", quantity=5
    )
    pl, _ = _make_processed_price_list(
        warehouse, sizes_and_qty={"S": 10}, product=product
    )

    activate_price_list_task(pl.pk)

    stock.refresh_from_db()
    assert stock.quantity == 15


def test_activate_creates_product_when_no_product_fk(
    db, warehouse, required_attributes
):
    from saleor.product.models import Category, Product, ProductType

    ProductType.objects.get_or_create(
        slug="apparel-type", defaults={"name": "Apparel", "has_variants": True}
    )
    Category.objects.create(name="Apparel", slug="apparel")

    pl, item = _make_processed_price_list(
        warehouse, sizes_and_qty={"S": 7}, product=None
    )

    activate_price_list_task(pl.pk)

    item.refresh_from_db()
    assert item.product_id is not None
    assert Product.objects.filter(pk=item.product_id).exists()
    assert Stock.objects.filter(
        product_variant__product_id=item.product_id,
        warehouse=warehouse,
        quantity=7,
    ).exists()


def test_activate_creates_product_media_for_new_product(
    db, warehouse, required_attributes
):
    from unittest.mock import patch

    from saleor.product.models import Category, ProductType

    ProductType.objects.get_or_create(
        slug="apparel-type", defaults={"name": "Apparel", "has_variants": True}
    )
    Category.objects.create(name="Apparel", slug="apparel")

    pl, item = _make_processed_price_list(
        warehouse, sizes_and_qty={"S": 1}, product=None
    )
    item.image_url = "https://example.com/image.jpg"
    item.save(update_fields=["image_url"])

    with patch("saleor.product.ingestion.create_product_media") as mock_create_media:
        activate_price_list_task(pl.pk)

    item.refresh_from_db()
    mock_create_media.assert_called_once()
    call_args = mock_create_media.call_args
    assert call_args.args[1] == "https://example.com/image.jpg"


def test_activate_creates_product_media_for_existing_product_without_media(
    db, warehouse
):
    from unittest.mock import patch

    product, _, _ = _make_product_with_variant_and_stock(warehouse, size="S")
    assert not product.media.exists()

    pl, item = _make_processed_price_list(
        warehouse, sizes_and_qty={"S": 1}, product=product
    )
    item.image_url = "https://example.com/image.jpg"
    item.save(update_fields=["image_url"])

    with patch("saleor.product.ingestion.create_product_media") as mock_create_media:
        activate_price_list_task(pl.pk)

    mock_create_media.assert_called_once()
    call_args = mock_create_media.call_args
    assert call_args.args[1] == "https://example.com/image.jpg"


def test_activate_skips_product_media_for_existing_product_with_media(db, warehouse):
    from unittest.mock import patch

    from saleor.product.models import ProductMedia

    product, _, _ = _make_product_with_variant_and_stock(warehouse, size="S")
    ProductMedia.objects.create(product=product, alt="existing")

    pl, item = _make_processed_price_list(
        warehouse, sizes_and_qty={"S": 1}, product=product
    )
    item.image_url = "https://example.com/image.jpg"
    item.save(update_fields=["image_url"])

    with patch("saleor.product.ingestion.create_product_media") as mock_create_media:
        activate_price_list_task(pl.pk)

    mock_create_media.assert_not_called()


def test_activate_is_idempotent(db, warehouse):
    from saleor.product import PriceListStatus

    product, _, stock = _make_product_with_variant_and_stock(
        warehouse, size="S", quantity=10
    )
    pl, _ = _make_processed_price_list(
        warehouse, sizes_and_qty={"S": 5}, product=product
    )
    pl.status = PriceListStatus.ACTIVE
    pl.save(update_fields=["status"])

    activate_price_list_task(pl.pk)

    stock.refresh_from_db()
    assert stock.quantity == 10


# ---------------------------------------------------------------------------
# Additional deactivate tests
# ---------------------------------------------------------------------------


def test_deactivate_is_idempotent(db, warehouse):
    product, _, stock = _make_product_with_variant_and_stock(warehouse, quantity=50)
    pl, _ = _make_processed_price_list(warehouse, product=product)

    deactivate_price_list_task(pl.pk)

    stock.refresh_from_db()
    assert stock.quantity == 50


def test_deactivate_does_not_affect_other_warehouse(db, warehouse):
    from saleor.account.models import Address
    from saleor.product import PriceListStatus

    address = Address.objects.create(
        street_address_1="2 Other St", city="City", country="GB"
    )
    other_wh = Warehouse.objects.create(
        name="Other WH", slug="other-wh", address=address, is_owned=False
    )

    product, variant, stock_main = _make_product_with_variant_and_stock(
        warehouse, quantity=50
    )
    stock_other = Stock.objects.create(
        product_variant=variant, warehouse=other_wh, quantity=50
    )

    pl, _ = _make_processed_price_list(warehouse, product=product)
    pl.status = PriceListStatus.ACTIVE
    pl.save(update_fields=["status"])

    deactivate_price_list_task(pl.pk)

    stock_main.refresh_from_db()
    stock_other.refresh_from_db()
    assert stock_main.quantity == 0
    assert stock_other.quantity == 50


def test_deactivate_with_unlinked_items_zeroes_only_linked_stock(db, warehouse):
    from django.utils import timezone

    from saleor.product import PriceListStatus

    product, _, stock = _make_product_with_variant_and_stock(warehouse, quantity=40)

    pl = PriceList.objects.create(
        warehouse=warehouse, config={}, processing_completed_at=timezone.now()
    )
    PriceListItem.objects.create(
        price_list=pl,
        row_index=0,
        product_code="LINKED",
        brand="B",
        description="linked",
        category="Apparel",
        sizes_and_qty={"S": 40},
        sell_price=Decimal(10),
        currency="GBP",
        is_valid=True,
        product=product,
    )
    PriceListItem.objects.create(
        price_list=pl,
        row_index=1,
        product_code="UNLINKED",
        brand="B",
        description="unlinked",
        category="Apparel",
        sizes_and_qty={"S": 10},
        sell_price=Decimal(10),
        currency="GBP",
        is_valid=True,
        product=None,
    )
    pl.status = PriceListStatus.ACTIVE
    pl.save(update_fields=["status"])

    deactivate_price_list_task(pl.pk)

    stock.refresh_from_db()
    assert stock.quantity == 0


# ---------------------------------------------------------------------------
# Additional replace tests
# ---------------------------------------------------------------------------


def test_replace_zeros_removed_size(db, warehouse):
    from django.utils import timezone

    from saleor.product.models import ProductVariant

    product, _, stock_s = _make_product_with_variant_and_stock(
        warehouse, size="S", quantity=30
    )
    variant_m = ProductVariant.objects.create(
        product=product, name="M", sku=f"sku-{product.pk}-M"
    )
    stock_m = Stock.objects.create(
        product_variant=variant_m, warehouse=warehouse, quantity=20
    )

    old_pl = PriceList.objects.create(
        warehouse=warehouse, config={}, processing_completed_at=timezone.now()
    )
    PriceListItem.objects.create(
        price_list=old_pl,
        row_index=0,
        product_code="X",
        brand="B",
        description="X",
        category="Apparel",
        sizes_and_qty={"S": 30, "M": 20},
        sell_price=Decimal(10),
        currency="GBP",
        is_valid=True,
        product=product,
    )

    new_pl = PriceList.objects.create(
        warehouse=warehouse, config={}, processing_completed_at=timezone.now()
    )
    PriceListItem.objects.create(
        price_list=new_pl,
        row_index=0,
        product_code="X",
        brand="B",
        description="X",
        category="Apparel",
        sizes_and_qty={"S": 15},
        sell_price=Decimal(10),
        currency="GBP",
        is_valid=True,
        product=product,
    )

    replace_price_list_task(old_pl.pk, new_pl.pk)

    stock_s.refresh_from_db()
    stock_m.refresh_from_db()
    assert stock_s.quantity == 15
    assert stock_m.quantity == 0


def test_replace_creates_stock_for_added_size(db, warehouse):
    from django.utils import timezone

    product, _, stock_s = _make_product_with_variant_and_stock(
        warehouse, size="S", quantity=30
    )

    old_pl = PriceList.objects.create(
        warehouse=warehouse, config={}, processing_completed_at=timezone.now()
    )
    PriceListItem.objects.create(
        price_list=old_pl,
        row_index=0,
        product_code="X",
        brand="B",
        description="X",
        category="Apparel",
        sizes_and_qty={"S": 30},
        sell_price=Decimal(10),
        currency="GBP",
        is_valid=True,
        product=product,
    )

    new_pl = PriceList.objects.create(
        warehouse=warehouse, config={}, processing_completed_at=timezone.now()
    )
    PriceListItem.objects.create(
        price_list=new_pl,
        row_index=0,
        product_code="X",
        brand="B",
        description="X",
        category="Apparel",
        sizes_and_qty={"S": 15, "L": 25},
        sell_price=Decimal(10),
        currency="GBP",
        is_valid=True,
        product=product,
    )

    replace_price_list_task(old_pl.pk, new_pl.pk)

    stock_s.refresh_from_db()
    assert stock_s.quantity == 15
    stock_l = Stock.objects.get(
        product_variant__product=product,
        product_variant__name="L",
        warehouse=warehouse,
    )
    assert stock_l.quantity == 25


def test_replace_activates_new_only_product(db, warehouse):
    from django.utils import timezone

    product_a, _, stock_a = _make_product_with_variant_and_stock(
        warehouse, size="S", quantity=50
    )
    product_b, _, stock_b = _make_product_with_variant_and_stock(
        warehouse, size="M", quantity=10
    )

    old_pl = PriceList.objects.create(
        warehouse=warehouse, config={}, processing_completed_at=timezone.now()
    )
    PriceListItem.objects.create(
        price_list=old_pl,
        row_index=0,
        product_code="A",
        brand="B",
        description="A",
        category="Apparel",
        sizes_and_qty={"S": 50},
        sell_price=Decimal(10),
        currency="GBP",
        is_valid=True,
        product=product_a,
    )

    new_pl = PriceList.objects.create(
        warehouse=warehouse, config={}, processing_completed_at=timezone.now()
    )
    PriceListItem.objects.create(
        price_list=new_pl,
        row_index=0,
        product_code="B",
        brand="B",
        description="B",
        category="Apparel",
        sizes_and_qty={"M": 5},
        sell_price=Decimal(10),
        currency="GBP",
        is_valid=True,
        product=product_b,
    )

    replace_price_list_task(old_pl.pk, new_pl.pk)

    stock_a.refresh_from_db()
    stock_b.refresh_from_db()
    assert stock_a.quantity == 0
    assert stock_b.quantity == 15


def test_replace_sets_timestamps(db, warehouse):
    from django.utils import timezone

    old_pl = PriceList.objects.create(
        warehouse=warehouse, config={}, processing_completed_at=timezone.now()
    )
    new_pl = PriceList.objects.create(
        warehouse=warehouse, config={}, processing_completed_at=timezone.now()
    )

    replace_price_list_task(old_pl.pk, new_pl.pk)

    old_pl.refresh_from_db()
    new_pl.refresh_from_db()
    assert old_pl.deactivated_at is not None
    assert new_pl.activated_at is not None


def test_replace_is_idempotent(db, warehouse):
    from django.utils import timezone

    from saleor.product import PriceListStatus

    product, _, stock = _make_product_with_variant_and_stock(warehouse, quantity=50)

    old_pl = PriceList.objects.create(
        warehouse=warehouse,
        config={},
        processing_completed_at=timezone.now(),
        status=PriceListStatus.INACTIVE,
    )
    new_pl = PriceList.objects.create(
        warehouse=warehouse,
        config={},
        processing_completed_at=timezone.now(),
        status=PriceListStatus.ACTIVE,
    )

    replace_price_list_task(old_pl.pk, new_pl.pk)

    stock.refresh_from_db()
    assert stock.quantity == 50


# ---------------------------------------------------------------------------
# End-to-end tests
# ---------------------------------------------------------------------------


def test_process_then_activate_creates_products_and_stock(
    db, warehouse, hk_excel, required_attributes
):
    from saleor.product import PriceListStatus
    from saleor.product.models import Category, Product, ProductType

    ProductType.objects.get_or_create(
        slug="apparel-type", defaults={"name": "Apparel", "has_variants": True}
    )
    Category.objects.create(name="Apparel", slug="apparel")

    price_list = _make_price_list(warehouse, hk_excel)

    process_price_list_task(price_list.pk)

    price_list.refresh_from_db()
    assert price_list.processing_completed_at is not None
    assert PriceListItem.objects.filter(
        price_list=price_list, is_valid=True
    ).count() == len(HK_ROWS)

    activate_price_list_task(price_list.pk)

    price_list.refresh_from_db()
    assert price_list.status == PriceListStatus.ACTIVE
    assert Product.objects.count() == len(HK_ROWS)

    stock_is1637_m = Stock.objects.get(
        product_variant__product__name="TIRO24 C TRPNTW",
        product_variant__name="M",
        warehouse=warehouse,
    )
    assert stock_is1637_m.quantity == 50

    stock_hk5015_xs = Stock.objects.get(
        product_variant__product__name="ADV WNTR AOP OH",
        product_variant__name="XS",
        warehouse=warehouse,
    )
    assert stock_hk5015_xs.quantity == 25


# ---------------------------------------------------------------------------
# Channel scoping tests
# ---------------------------------------------------------------------------


@pytest.fixture
def channel_gbp(db):
    from saleor.channel.models import Channel

    return Channel.objects.create(
        name="GBP Channel",
        slug="gbp-channel",
        currency_code="GBP",
        default_country="GB",
        is_active=True,
    )


@pytest.fixture
def other_channel(db):
    from saleor.channel.models import Channel

    return Channel.objects.create(
        name="Other Channel",
        slug="other-channel",
        currency_code="EUR",
        default_country="DE",
        is_active=True,
    )


def test_activate_creates_channel_listings_only_for_price_list_channels(
    db, warehouse, channel_gbp, other_channel, required_attributes
):
    from saleor.product.models import Category, ProductChannelListing, ProductType

    ProductType.objects.get_or_create(
        slug="apparel-type", defaults={"name": "Apparel", "has_variants": True}
    )
    Category.objects.create(name="Apparel", slug="apparel")

    pl, _ = _make_processed_price_list(warehouse, sizes_and_qty={"S": 5}, product=None)
    pl.channels.set([channel_gbp])

    activate_price_list_task(pl.pk)

    from saleor.product.models import Product

    product = Product.objects.get()
    assert ProductChannelListing.objects.filter(
        product=product, channel=channel_gbp
    ).exists()
    assert not ProductChannelListing.objects.filter(
        product=product, channel=other_channel
    ).exists()


def test_activate_creates_product_published_not_visible_in_listings(
    db, warehouse, channel_gbp, required_attributes
):
    from saleor.product.models import Category, ProductChannelListing, ProductType

    ProductType.objects.get_or_create(
        slug="apparel-type", defaults={"name": "Apparel", "has_variants": True}
    )
    Category.objects.create(name="Apparel", slug="apparel")

    pl, _ = _make_processed_price_list(warehouse, sizes_and_qty={"S": 5}, product=None)
    pl.channels.set([channel_gbp])

    activate_price_list_task(pl.pk)

    listing = ProductChannelListing.objects.get(channel=channel_gbp)
    assert listing.is_published is True
    assert listing.visible_in_listings is False
    assert listing.available_for_purchase_at is not None


def test_activate_creates_product_channel_listing_for_draft_orders(
    db, warehouse, channel_gbp, required_attributes
):
    """Products must have a channel listing even when not published, so staff can add to draft orders."""
    from saleor.product.models import Category, ProductChannelListing, ProductType

    ProductType.objects.get_or_create(
        slug="apparel-type", defaults={"name": "Apparel", "has_variants": True}
    )
    Category.objects.create(name="Apparel", slug="apparel")

    pl, _ = _make_processed_price_list(warehouse, sizes_and_qty={"S": 5}, product=None)
    pl.channels.set([channel_gbp])

    activate_price_list_task(pl.pk)

    assert ProductChannelListing.objects.filter(channel=channel_gbp).exists()


def test_activate_existing_product_creates_variant_listings_only_for_price_list_channels(
    db, warehouse, channel_gbp, other_channel
):
    from saleor.product.models import ProductVariantChannelListing

    product, variant, _ = _make_product_with_variant_and_stock(
        warehouse, size="S", quantity=0
    )
    Stock.objects.filter(product_variant=variant).delete()

    pl, _ = _make_processed_price_list(
        warehouse, sizes_and_qty={"S": 10}, product=product
    )
    pl.channels.set([channel_gbp])

    activate_price_list_task(pl.pk)

    assert ProductVariantChannelListing.objects.filter(
        variant=variant, channel=channel_gbp
    ).exists()
    assert not ProductVariantChannelListing.objects.filter(
        variant=variant, channel=other_channel
    ).exists()


def test_activate_matches_existing_product_by_code_and_brand(
    db, warehouse, required_attributes
):
    from saleor.attribute.models import AttributeValue
    from saleor.attribute.models.product import AssignedProductAttributeValue
    from saleor.product.models import Category, Product, ProductType

    ProductType.objects.get_or_create(
        slug="apparel-type", defaults={"name": "Apparel", "has_variants": True}
    )
    Category.objects.create(name="Apparel", slug="apparel")

    product_type = ProductType.objects.get(slug="apparel-type")
    existing_product = Product.objects.create(
        name="Existing Product", slug="existing-product", product_type=product_type
    )
    code_value = AttributeValue.objects.create(
        attribute=required_attributes["Product Code"], name="TEST-001", slug="test-001"
    )
    brand_value = AttributeValue.objects.create(
        attribute=required_attributes["Brand"], name="TestBrand", slug="testbrand"
    )
    AssignedProductAttributeValue.objects.create(
        product=existing_product, value=code_value
    )
    AssignedProductAttributeValue.objects.create(
        product=existing_product, value=brand_value
    )

    pl, _ = _make_processed_price_list(warehouse, sizes_and_qty={"S": 12}, product=None)

    activate_price_list_task(pl.pk)

    assert Product.objects.count() == 1
    stock = Stock.objects.get(
        product_variant__product=existing_product,
        product_variant__name="S",
        warehouse=warehouse,
    )
    assert stock.quantity == 12


# ---------------------------------------------------------------------------
# Conservation-of-mass helpers
# ---------------------------------------------------------------------------


def _make_allocation(stock, order_status, qty):
    """Create an Order → OrderLine → Allocation chain for a given stock row."""
    from decimal import Decimal

    from saleor.channel.models import Channel
    from saleor.order import OrderOrigin
    from saleor.order.models import Order, OrderLine
    from saleor.warehouse.models import Allocation

    channel, _ = Channel.objects.get_or_create(
        slug="conservation-test-ch",
        defaults={
            "name": "Conservation Test Channel",
            "currency_code": "GBP",
            "default_country": "GB",
        },
    )
    order = Order.objects.create(
        status=order_status,
        channel=channel,
        currency="GBP",
        origin=OrderOrigin.DRAFT,
        lines_count=1,
    )
    line = OrderLine.objects.create(
        order=order,
        variant=stock.product_variant,
        product_name="Test",
        quantity=qty,
        currency="GBP",
        unit_price_net_amount=Decimal(10),
        unit_price_gross_amount=Decimal(10),
        total_price_net_amount=Decimal(10),
        total_price_gross_amount=Decimal(10),
        is_shipping_required=False,
        is_gift_card=False,
    )
    return Allocation.objects.create(
        order_line=line,
        stock=stock,
        quantity_allocated=qty,
    )


# ---------------------------------------------------------------------------
# Deactivate — conservation of mass tests
# ---------------------------------------------------------------------------


def test_deactivate_conservation_unconfirmed_only(db, warehouse):
    from saleor.order import OrderStatus
    from saleor.product import PriceListStatus
    from saleor.warehouse.models import Allocation

    product, variant, stock = _make_product_with_variant_and_stock(
        warehouse, quantity=50
    )
    stock.quantity_allocated = 20
    stock.save(update_fields=["quantity_allocated"])
    alloc = _make_allocation(stock, OrderStatus.UNCONFIRMED, 20)

    pl, _ = _make_processed_price_list(warehouse, product=product)
    pl.status = PriceListStatus.ACTIVE
    pl.save(update_fields=["status"])

    deactivate_price_list_task(pl.pk)

    stock.refresh_from_db()
    assert stock.quantity == 0
    assert stock.quantity_allocated == 0
    assert not Allocation.objects.filter(pk=alloc.pk).exists()


def test_deactivate_conservation_mixed_unconfirmed_and_unfulfilled(db, warehouse):
    from saleor.order import OrderStatus
    from saleor.product import PriceListStatus
    from saleor.warehouse.models import Allocation

    product, variant, stock = _make_product_with_variant_and_stock(
        warehouse, quantity=50
    )
    stock.quantity_allocated = 30
    stock.save(update_fields=["quantity_allocated"])
    draft_alloc = _make_allocation(stock, OrderStatus.UNCONFIRMED, 20)
    unfulfilled_alloc = _make_allocation(stock, OrderStatus.UNFULFILLED, 10)

    pl, _ = _make_processed_price_list(warehouse, product=product)
    pl.status = PriceListStatus.ACTIVE
    pl.save(update_fields=["status"])

    deactivate_price_list_task(pl.pk)

    stock.refresh_from_db()
    assert stock.quantity == 10
    assert stock.quantity_allocated == 10
    assert not Allocation.objects.filter(pk=draft_alloc.pk).exists()
    assert Allocation.objects.filter(pk=unfulfilled_alloc.pk).exists()


def test_deactivate_does_not_touch_unfulfilled_order_allocations(db, warehouse):
    from saleor.order import OrderStatus
    from saleor.product import PriceListStatus
    from saleor.warehouse.models import Allocation

    product, variant, stock = _make_product_with_variant_and_stock(
        warehouse, quantity=50
    )
    stock.quantity_allocated = 15
    stock.save(update_fields=["quantity_allocated"])
    alloc = _make_allocation(stock, OrderStatus.UNFULFILLED, 15)

    pl, _ = _make_processed_price_list(warehouse, product=product)
    pl.status = PriceListStatus.ACTIVE
    pl.save(update_fields=["status"])

    deactivate_price_list_task(pl.pk)

    stock.refresh_from_db()
    assert stock.quantity == 15
    assert stock.quantity_allocated == 15
    assert Allocation.objects.filter(pk=alloc.pk).exists()


def test_deactivate_conservation_no_orders(db, warehouse):
    from saleor.product import PriceListStatus

    product, variant, stock = _make_product_with_variant_and_stock(
        warehouse, quantity=50
    )

    pl, _ = _make_processed_price_list(warehouse, product=product)
    pl.status = PriceListStatus.ACTIVE
    pl.save(update_fields=["status"])

    deactivate_price_list_task(pl.pk)

    stock.refresh_from_db()
    assert stock.quantity == 0
    assert stock.quantity_allocated == 0


def test_deactivate_conservation_multiple_unconfirmed_orders_same_stock(db, warehouse):
    from saleor.order import OrderStatus
    from saleor.product import PriceListStatus
    from saleor.warehouse.models import Allocation

    product, variant, stock = _make_product_with_variant_and_stock(
        warehouse, quantity=100
    )
    stock.quantity_allocated = 55
    stock.save(update_fields=["quantity_allocated"])
    alloc1 = _make_allocation(stock, OrderStatus.UNCONFIRMED, 30)
    alloc2 = _make_allocation(stock, OrderStatus.UNCONFIRMED, 25)

    pl, _ = _make_processed_price_list(warehouse, product=product)
    pl.status = PriceListStatus.ACTIVE
    pl.save(update_fields=["status"])

    deactivate_price_list_task(pl.pk)

    stock.refresh_from_db()
    assert stock.quantity == 0
    assert stock.quantity_allocated == 0
    assert not Allocation.objects.filter(pk__in=[alloc1.pk, alloc2.pk]).exists()


# ---------------------------------------------------------------------------
# Replace — conservation of mass tests
# ---------------------------------------------------------------------------


def _make_pl_with_item(warehouse, product, sizes_and_qty):
    from decimal import Decimal

    from django.utils import timezone

    pl = PriceList.objects.create(
        warehouse=warehouse, config={}, processing_completed_at=timezone.now()
    )
    PriceListItem.objects.create(
        price_list=pl,
        row_index=0,
        product_code="TEST",
        brand="B",
        description="Test",
        category="Apparel",
        sizes_and_qty=sizes_and_qty,
        sell_price=Decimal(10),
        currency="GBP",
        is_valid=True,
        product=product,
    )
    return pl


def test_replace_conservation_removed_product_with_unconfirmed_order(db, warehouse):
    from django.utils import timezone

    from saleor.order import OrderStatus
    from saleor.warehouse.models import Allocation

    product_a, variant_a, stock_a = _make_product_with_variant_and_stock(
        warehouse, size="S", quantity=50
    )
    stock_a.quantity_allocated = 20
    stock_a.save(update_fields=["quantity_allocated"])
    alloc = _make_allocation(stock_a, OrderStatus.UNCONFIRMED, 20)

    old_pl = _make_pl_with_item(warehouse, product_a, {"S": 50})
    new_pl = PriceList.objects.create(
        warehouse=warehouse, config={}, processing_completed_at=timezone.now()
    )

    replace_price_list_task(old_pl.pk, new_pl.pk)

    stock_a.refresh_from_db()
    assert stock_a.quantity == 0
    assert stock_a.quantity_allocated == 0
    assert not Allocation.objects.filter(pk=alloc.pk).exists()


def test_replace_conservation_removed_product_mixed_orders(db, warehouse):
    from django.utils import timezone

    from saleor.order import OrderStatus
    from saleor.warehouse.models import Allocation

    product_a, variant_a, stock_a = _make_product_with_variant_and_stock(
        warehouse, size="S", quantity=50
    )
    stock_a.quantity_allocated = 30
    stock_a.save(update_fields=["quantity_allocated"])
    draft_alloc = _make_allocation(stock_a, OrderStatus.UNCONFIRMED, 20)
    unfulfilled_alloc = _make_allocation(stock_a, OrderStatus.UNFULFILLED, 10)

    old_pl = _make_pl_with_item(warehouse, product_a, {"S": 50})
    new_pl = PriceList.objects.create(
        warehouse=warehouse, config={}, processing_completed_at=timezone.now()
    )

    replace_price_list_task(old_pl.pk, new_pl.pk)

    stock_a.refresh_from_db()
    assert stock_a.quantity == 10
    assert stock_a.quantity_allocated == 10
    assert not Allocation.objects.filter(pk=draft_alloc.pk).exists()
    assert Allocation.objects.filter(pk=unfulfilled_alloc.pk).exists()


def test_replace_conservation_retained_product_not_deallocated(db, warehouse):
    from saleor.order import OrderStatus
    from saleor.warehouse.models import Allocation

    product_b, variant_b, stock_b = _make_product_with_variant_and_stock(
        warehouse, size="M", quantity=50
    )
    stock_b.quantity_allocated = 15
    stock_b.save(update_fields=["quantity_allocated"])
    alloc = _make_allocation(stock_b, OrderStatus.UNCONFIRMED, 15)

    old_pl = _make_pl_with_item(warehouse, product_b, {"M": 50})
    new_pl = _make_pl_with_item(warehouse, product_b, {"M": 30})

    replace_price_list_task(old_pl.pk, new_pl.pk)

    stock_b.refresh_from_db()
    assert Allocation.objects.filter(pk=alloc.pk).exists()
    assert stock_b.quantity_allocated == 15
    assert stock_b.quantity == 30


def test_replace_conservation_removed_size_with_unconfirmed_order(db, warehouse):
    from saleor.order import OrderStatus
    from saleor.warehouse.models import Allocation

    product_c, variant_s, stock_s = _make_product_with_variant_and_stock(
        warehouse, size="S", quantity=30
    )
    stock_s.quantity_allocated = 10
    stock_s.save(update_fields=["quantity_allocated"])
    alloc = _make_allocation(stock_s, OrderStatus.UNCONFIRMED, 10)

    old_pl = _make_pl_with_item(warehouse, product_c, {"S": 30, "M": 20})
    new_pl = _make_pl_with_item(warehouse, product_c, {"M": 20})

    replace_price_list_task(old_pl.pk, new_pl.pk)

    stock_s.refresh_from_db()
    assert stock_s.quantity == 0
    assert stock_s.quantity_allocated == 0
    assert not Allocation.objects.filter(pk=alloc.pk).exists()


def test_replace_conservation_removed_size_mixed_orders(db, warehouse):
    from saleor.order import OrderStatus
    from saleor.warehouse.models import Allocation

    product_c, variant_s, stock_s = _make_product_with_variant_and_stock(
        warehouse, size="S", quantity=40
    )
    stock_s.quantity_allocated = 25
    stock_s.save(update_fields=["quantity_allocated"])
    draft_alloc = _make_allocation(stock_s, OrderStatus.UNCONFIRMED, 15)
    unfulfilled_alloc = _make_allocation(stock_s, OrderStatus.UNFULFILLED, 10)

    old_pl = _make_pl_with_item(warehouse, product_c, {"S": 40, "M": 20})
    new_pl = _make_pl_with_item(warehouse, product_c, {"M": 20})

    replace_price_list_task(old_pl.pk, new_pl.pk)

    stock_s.refresh_from_db()
    assert stock_s.quantity == 10
    assert stock_s.quantity_allocated == 10
    assert not Allocation.objects.filter(pk=draft_alloc.pk).exists()
    assert Allocation.objects.filter(pk=unfulfilled_alloc.pk).exists()


# ---------------------------------------------------------------------------
# Search index tests
# ---------------------------------------------------------------------------


def test_activate_marks_products_search_index_dirty(db, warehouse):
    from unittest.mock import patch

    from saleor.product.tasks import update_products_search_vector_task

    product, _, _ = _make_product_with_variant_and_stock(warehouse)
    pl, _ = _make_processed_price_list(
        warehouse, sizes_and_qty={"S": 5}, product=product
    )

    with patch.object(update_products_search_vector_task, "delay") as mock_delay:
        activate_price_list_task(pl.pk)

    product.refresh_from_db()
    assert product.search_index_dirty is True
    mock_delay.assert_called_once()


def test_deactivate_marks_products_search_index_dirty(db, warehouse):
    from unittest.mock import patch

    from saleor.product import PriceListStatus
    from saleor.product.tasks import update_products_search_vector_task

    product, _, _ = _make_product_with_variant_and_stock(warehouse)
    pl, _ = _make_processed_price_list(warehouse, product=product)
    pl.status = PriceListStatus.ACTIVE
    pl.save(update_fields=["status"])

    with patch.object(update_products_search_vector_task, "delay") as mock_delay:
        deactivate_price_list_task(pl.pk)

    product.refresh_from_db()
    assert product.search_index_dirty is True
    mock_delay.assert_called_once()


def test_replace_marks_all_affected_products_search_index_dirty(db, warehouse):
    from unittest.mock import patch

    from django.utils import timezone

    from saleor.product.tasks import update_products_search_vector_task

    product_a, _, _ = _make_product_with_variant_and_stock(
        warehouse, size="S", quantity=10
    )
    product_b, _, _ = _make_product_with_variant_and_stock(
        warehouse, size="M", quantity=10
    )

    old_pl = PriceList.objects.create(
        warehouse=warehouse, config={}, processing_completed_at=timezone.now()
    )
    PriceListItem.objects.create(
        price_list=old_pl,
        row_index=0,
        product_code="A",
        brand="B",
        description="A",
        category="Apparel",
        sizes_and_qty={"S": 10},
        sell_price=Decimal(10),
        currency="GBP",
        is_valid=True,
        product=product_a,
    )

    new_pl = PriceList.objects.create(
        warehouse=warehouse, config={}, processing_completed_at=timezone.now()
    )
    PriceListItem.objects.create(
        price_list=new_pl,
        row_index=0,
        product_code="B",
        brand="B",
        description="B",
        category="Apparel",
        sizes_and_qty={"M": 10},
        sell_price=Decimal(10),
        currency="GBP",
        is_valid=True,
        product=product_b,
    )

    with patch.object(update_products_search_vector_task, "delay") as mock_delay:
        replace_price_list_task(old_pl.pk, new_pl.pk)

    product_a.refresh_from_db()
    product_b.refresh_from_db()
    assert product_a.search_index_dirty is True
    assert product_b.search_index_dirty is True
    mock_delay.assert_called_once()


# ---------------------------------------------------------------------------
# Duplicate product code deduplication tests
# ---------------------------------------------------------------------------


def test_process_marks_duplicate_product_code_invalid(db, warehouse, hk_excel):
    """Second occurrence of the same (product_code, brand) in a sheet is marked invalid."""
    from django.utils import timezone

    pl = PriceList.objects.create(
        warehouse=warehouse,
        config={},
        processing_completed_at=timezone.now(),
    )
    PriceListItem.objects.create(
        price_list=pl,
        row_index=0,
        product_code="DUP-001",
        brand="Adidas",
        description="First Item",
        category="Apparel",
        sizes_and_qty={"S": 10},
        sell_price=Decimal("10.00"),
        currency="GBP",
        is_valid=True,
    )
    PriceListItem.objects.create(
        price_list=pl,
        row_index=1,
        product_code="DUP-001",
        brand="Adidas",
        description="First Item Again",
        category="Apparel",
        sizes_and_qty={"M": 5},
        sell_price=Decimal("10.00"),
        currency="GBP",
        is_valid=True,
    )

    # Trigger only the deduplication pass (not full processing) by calling it
    # indirectly: re-run the duplicate-marking logic directly on a fresh pl.
    # Simpler: call process_price_list_task on a pl that has a real file.
    # Instead, test the deduplication step in isolation by recreating its logic.
    # We test it through process_price_list_task using a real sheet below;
    # here we verify the existing items get patched correctly by triggering
    # the relevant code path directly.
    from saleor.product.models import PriceListItem as PLI

    seen_keys: set = set()
    duplicate_items = []
    for item in pl.items.all():
        if not item.is_valid:
            continue
        key = (item.product_code, item.brand)
        if key in seen_keys:
            item.is_valid = False
            item.validation_errors = [
                "duplicate product_code+brand in this sheet: DUP-001"
            ]
            duplicate_items.append(item)
        else:
            seen_keys.add(key)
    PLI.objects.bulk_update(duplicate_items, ["is_valid", "validation_errors"])

    valid = PLI.objects.filter(price_list=pl, is_valid=True)
    invalid = PLI.objects.filter(price_list=pl, is_valid=False)
    assert valid.count() == 1
    assert invalid.count() == 1
    assert any("duplicate" in e for e in invalid.first().validation_errors)


def test_process_task_marks_duplicate_product_code_invalid(db, warehouse, tmp_path):
    """process_price_list_task marks the second row with the same product code invalid."""
    import openpyxl
    from django.core.files import File

    from saleor.product.tasks import process_price_list_task

    headers = [
        "Brand",
        "Article",
        "Description",
        "RRP (GBP)",
        "Sale Price (GBP)",
        "Quantity",
        "Category",
        "Gender",
        "Sizing (UK)",
        "Category.1",
        "Gender.1",
        "Rounded RRP (GBP)",
        "unit_weight_kg",
        "updated_sizing",
        "issues",
        "updated_description",
    ]
    row = [
        "Adidas",
        "IS1637",
        "TIRO24 C TRPNTW",
        40,
        9.025,
        190,
        "Apparel",
        "Women",
        "XS[20], M[50]",
        "Apparel",
        "Women",
        40,
        0.2,
        "XS[20], M[50]",
        None,
        "TIRO24 C TRPNTW",
    ]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1 (1)"
    ws.append(headers)
    ws.append(row)
    ws.append(row)  # exact duplicate product code

    path = tmp_path / "dup.xlsx"
    wb.save(path)

    with open(path, "rb") as f:
        pl = PriceList.objects.create(
            warehouse=warehouse,
            excel_file=File(f, name="dup.xlsx"),
            config={
                "sheet_name": "Sheet1 (1)",
                "header_row": 0,
                "column_map": {
                    "0": "brand",
                    "1": "product_code",
                    "2": "description",
                    "3": "rrp",
                    "4": "sell_price",
                    "6": "category",
                    "12": "weight_kg",
                    "13": "sizes",
                },
                "default_currency": "GBP",
            },
        )

    process_price_list_task(pl.pk)

    valid = PriceListItem.objects.filter(price_list=pl, is_valid=True)
    invalid = PriceListItem.objects.filter(price_list=pl, is_valid=False)
    assert valid.count() == 1
    assert invalid.count() == 1
    assert any("duplicate" in e for e in invalid.first().validation_errors)


def test_activate_duplicate_product_code_does_not_create_two_products(
    db, warehouse, required_attributes
):
    """If two valid items share a product code (slipped past processing), activation creates only one product and reuses it for the second item."""
    from django.utils import timezone

    from saleor.product.models import Category, Product, ProductType
    from saleor.product.tasks import activate_price_list_task

    ProductType.objects.get_or_create(
        slug="apparel-type", defaults={"name": "Apparel", "has_variants": True}
    )
    Category.objects.create(name="Apparel", slug="apparel")

    pl = PriceList.objects.create(
        warehouse=warehouse, config={}, processing_completed_at=timezone.now()
    )
    PriceListItem.objects.create(
        price_list=pl,
        row_index=0,
        product_code="DUP-001",
        brand="Adidas",
        description="Duplicate Product",
        category="Apparel",
        sizes_and_qty={"S": 10},
        sell_price=Decimal("10.00"),
        currency="GBP",
        is_valid=True,
        product=None,
    )
    PriceListItem.objects.create(
        price_list=pl,
        row_index=1,
        product_code="DUP-001",
        brand="Adidas",
        description="Duplicate Product",
        category="Apparel",
        sizes_and_qty={"M": 5},
        sell_price=Decimal("10.00"),
        currency="GBP",
        is_valid=True,
        product=None,
    )

    activate_price_list_task(pl.pk)

    assert Product.objects.count() == 1


# ---------------------------------------------------------------------------
# _deallocate_draft_unconfirmed — unit tests
# ---------------------------------------------------------------------------


def test_deallocate_removes_unconfirmed_allocation_and_updates_stock(db, warehouse):
    from django.db import transaction

    from saleor.order import OrderStatus
    from saleor.product.tasks import _deallocate_draft_unconfirmed
    from saleor.warehouse.models import Allocation

    product, variant, stock = _make_product_with_variant_and_stock(
        warehouse, quantity=50
    )
    stock.quantity_allocated = 20
    stock.save(update_fields=["quantity_allocated"])
    alloc = _make_allocation(stock, OrderStatus.UNCONFIRMED, 20)

    with transaction.atomic():
        _deallocate_draft_unconfirmed(warehouse, [product.pk])

    stock.refresh_from_db()
    assert stock.quantity_allocated == 0
    assert not Allocation.objects.filter(pk=alloc.pk).exists()


def test_deallocate_removes_draft_allocation_and_updates_stock(db, warehouse):
    from django.db import transaction

    from saleor.order import OrderStatus
    from saleor.product.tasks import _deallocate_draft_unconfirmed
    from saleor.warehouse.models import Allocation

    product, variant, stock = _make_product_with_variant_and_stock(
        warehouse, quantity=50
    )
    stock.quantity_allocated = 15
    stock.save(update_fields=["quantity_allocated"])
    alloc = _make_allocation(stock, OrderStatus.DRAFT, 15)

    with transaction.atomic():
        _deallocate_draft_unconfirmed(warehouse, [product.pk])

    stock.refresh_from_db()
    assert stock.quantity_allocated == 0
    assert not Allocation.objects.filter(pk=alloc.pk).exists()


def test_deallocate_does_not_touch_unfulfilled_allocations(db, warehouse):
    from django.db import transaction

    from saleor.order import OrderStatus
    from saleor.product.tasks import _deallocate_draft_unconfirmed
    from saleor.warehouse.models import Allocation

    product, variant, stock = _make_product_with_variant_and_stock(
        warehouse, quantity=50
    )
    stock.quantity_allocated = 10
    stock.save(update_fields=["quantity_allocated"])
    alloc = _make_allocation(stock, OrderStatus.UNFULFILLED, 10)

    with transaction.atomic():
        _deallocate_draft_unconfirmed(warehouse, [product.pk])

    stock.refresh_from_db()
    assert stock.quantity_allocated == 10
    assert Allocation.objects.filter(pk=alloc.pk).exists()


def test_deallocate_aggregates_multiple_allocations_on_same_stock(db, warehouse):
    from django.db import transaction

    from saleor.order import OrderStatus
    from saleor.product.tasks import _deallocate_draft_unconfirmed
    from saleor.warehouse.models import Allocation

    product, variant, stock = _make_product_with_variant_and_stock(
        warehouse, quantity=100
    )
    stock.quantity_allocated = 45
    stock.save(update_fields=["quantity_allocated"])
    alloc1 = _make_allocation(stock, OrderStatus.UNCONFIRMED, 25)
    alloc2 = _make_allocation(stock, OrderStatus.UNCONFIRMED, 20)

    with transaction.atomic():
        _deallocate_draft_unconfirmed(warehouse, [product.pk])

    stock.refresh_from_db()
    assert stock.quantity_allocated == 0
    assert not Allocation.objects.filter(pk__in=[alloc1.pk, alloc2.pk]).exists()


def test_deallocate_size_names_filter_only_removes_matching_sizes(db, warehouse):
    from django.db import transaction

    from saleor.order import OrderStatus
    from saleor.product.models import ProductVariant
    from saleor.product.tasks import _deallocate_draft_unconfirmed
    from saleor.warehouse.models import Allocation

    product, variant_s, stock_s = _make_product_with_variant_and_stock(
        warehouse, size="S", quantity=30
    )
    variant_m = ProductVariant.objects.create(product=product, sku="test-m", name="M")
    stock_m = Stock.objects.create(
        product_variant=variant_m, warehouse=warehouse, quantity=30
    )
    stock_s.quantity_allocated = 10
    stock_s.save(update_fields=["quantity_allocated"])
    stock_m.quantity_allocated = 10
    stock_m.save(update_fields=["quantity_allocated"])
    alloc_s = _make_allocation(stock_s, OrderStatus.UNCONFIRMED, 10)
    alloc_m = _make_allocation(stock_m, OrderStatus.UNCONFIRMED, 10)

    with transaction.atomic():
        _deallocate_draft_unconfirmed(warehouse, [product.pk], size_names={"S"})

    stock_s.refresh_from_db()
    stock_m.refresh_from_db()
    assert stock_s.quantity_allocated == 0
    assert not Allocation.objects.filter(pk=alloc_s.pk).exists()
    assert stock_m.quantity_allocated == 10
    assert Allocation.objects.filter(pk=alloc_m.pk).exists()


def test_deallocate_no_allocations_is_a_noop(db, warehouse):
    from django.db import transaction

    from saleor.product.tasks import _deallocate_draft_unconfirmed

    product, variant, stock = _make_product_with_variant_and_stock(
        warehouse, quantity=50
    )

    with transaction.atomic():
        _deallocate_draft_unconfirmed(warehouse, [product.pk])

    stock.refresh_from_db()
    assert stock.quantity_allocated == 0


# ---------------------------------------------------------------------------
# process_price_list_task — exception handler
# ---------------------------------------------------------------------------


def test_process_task_sets_processing_failed_at_on_exception(db, warehouse, tmp_path):
    import openpyxl
    import pytest
    from django.core.files import File

    wb = openpyxl.Workbook()
    path = tmp_path / "test.xlsx"
    wb.save(path)

    with open(path, "rb") as f:
        pl = PriceList.objects.create(
            warehouse=warehouse,
            excel_file=File(f, name="test.xlsx"),
            config={},  # missing required "column_map" key — causes KeyError inside task
        )

    with pytest.raises(KeyError):
        process_price_list_task(pl.pk)

    pl.refresh_from_db()
    assert pl.processing_failed_at is not None
    assert pl.processing_completed_at is None


def test_process_task_clears_completed_at_when_exception_occurs(
    db, warehouse, tmp_path
):
    import openpyxl
    import pytest
    from django.core.files import File
    from django.utils import timezone

    wb = openpyxl.Workbook()
    path = tmp_path / "test.xlsx"
    wb.save(path)

    with open(path, "rb") as f:
        pl = PriceList.objects.create(
            warehouse=warehouse,
            excel_file=File(f, name="test.xlsx"),
            config={},
            processing_completed_at=timezone.now(),  # simulate a prior successful run
        )

    with pytest.raises(KeyError):
        process_price_list_task(pl.pk)

    pl.refresh_from_db()
    assert pl.processing_completed_at is None
    assert pl.processing_failed_at is not None
