import logging
import re
from decimal import Decimal

import attrs
import openpyxl
import pandas as pd
from django.core.files.base import ContentFile
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from django.utils.text import slugify

from saleor.attribute.models import Attribute, AttributeValue
from saleor.attribute.models.product import (
    AssignedProductAttributeValue,
)
from saleor.attribute.models.product_variant import (
    AssignedVariantAttribute,
    AssignedVariantAttributeValue,
    AttributeVariant,
)
from saleor.channel.models import Channel
from saleor.core.http_client import HTTPClient
from saleor.warehouse.models import Stock, Warehouse

from ...models import (
    Category,
    ProductChannelListing,
    ProductMedia,
    ProductType,
    ProductVariant,
    ProductVariantChannelListing,
)
from ...models import Product as ProductModel
from ...utils.variant_prices import update_discounted_prices_for_promotion

logger = logging.getLogger(__name__)


@attrs.define
class ProductData:
    product_code: str
    description: str
    category: str
    sizes: tuple[str]
    qty: tuple[int]
    brand: str
    rrp: float
    price: float
    currency: str
    image_data: dict | None = None  # Store image data with product


class Command(BaseCommand):
    help = "Ingest products from an Excel file."

    def add_arguments(self, parser):
        parser.add_argument(
            "excel_file",
            type=str,
            help="Path to the Excel file containing product data",
        )
        parser.add_argument(
            "--dry-run",
            action="store_true",
            help="Run without making any changes to the database",
        )
        parser.add_argument(
            "--sheet",
            type=str,
            default="Sheet1",
            help="Name of the sheet to read from (default: Sheet1)",
        )

    def handle(self, *args, **options):
        excel_file = options["excel_file"]
        dry_run = options["dry_run"]
        sheet_name = options["sheet"]

        self.stdout.write(f"Reading products from: {excel_file}")

        if dry_run:
            self.stdout.write(
                self.style.WARNING("Running in dry-run mode - no changes will be saved")
            )

        try:
            # Check available sheets first
            workbook = openpyxl.load_workbook(excel_file)
            available_sheets = workbook.sheetnames
            self.stdout.write(f"Available sheets: {', '.join(available_sheets)}")

            # Use first sheet if the specified one doesn't exist
            if sheet_name not in available_sheets:
                sheet_name = available_sheets[0]
                self.stdout.write(f"Using sheet: '{sheet_name}'")

            # Extract images from Excel using openpyxl
            images_by_row = self._extract_images(excel_file, sheet_name)

            # Read data using pandas
            df = pd.read_excel(excel_file, sheet_name=sheet_name)

            # Expected columns: Image, Code, Description, Category, Sizes, Brand, Qty, RRP, Price
            self.stdout.write(self.style.SUCCESS("\n=== Summary ==="))
            self.stdout.write(f"Total rows in Excel: {len(df)}")
            self.stdout.write(f"Total images extracted: {len(images_by_row)}")
            self.stdout.write(f"Columns found: {', '.join(df.columns.tolist())}")

            # Show image distribution
            if images_by_row:
                self.stdout.write(
                    f"\nImages found at rows: {sorted(images_by_row.keys())}"
                )

            # Process each row
            self.stdout.write("\n=== Processing Products ===")
            products = []
            for idx, row in df.iterrows():
                image_data = images_by_row.get(
                    idx + 2
                )  # +2 because pandas is 0-indexed and Excel has header row
                product_data = self._process_row(row, image_data)
                if product_data:
                    products.append(product_data)

            # Display first 3 products
            self.stdout.write("\n=== First 3 Products (parsed) ===")
            for product in products[:3]:
                self.stdout.write(f"\nProduct: {product.description}")
                self.stdout.write(f"  Code: {product.product_code}")
                self.stdout.write(f"  Brand: {product.brand}")
                self.stdout.write(f"  Category: {product.category}")
                self.stdout.write(f"  RRP: {product.rrp} {product.currency}")
                self.stdout.write(f"  Price: {product.price} {product.currency}")
                self.stdout.write(f"  Variants ({len(product.sizes)}):")
                for size, qty in zip(product.sizes, product.qty, strict=False):
                    self.stdout.write(f"    - Size {size}: {qty} in stock")

            self.stdout.write("\n=== Summary ===")
            self.stdout.write(f"Total products parsed: {len(products)}")

            # De-duplicate products by product code
            products = self._deduplicate_products(products)
            self.stdout.write(f"After de-duplication: {len(products)} products")

            if dry_run:
                self.stdout.write(
                    self.style.WARNING(
                        "\n=== DRY RUN - No database operations performed ==="
                    )
                )
                return

            # Validation phase
            self.stdout.write("\n=== Validation Phase ===")
            product_type_map = self._validate_product_types(products)
            category_map = self._validate_categories(products)
            attribute_map = self._validate_attributes()
            moq_value = self._get_minimum_order_quantity()
            self._validate_product_type_attributes(product_type_map, attribute_map)
            self._validate_sizes(products, attribute_map["Size"])
            self._validate_unique_product_codes(products)
            self._validate_unique_names(products)
            channels = self._validate_and_fetch_channels()
            warehouse = self._select_warehouse()
            exchange_rates = self._fetch_exchange_rates(products, channels)
            self._confirm_price_interpretation()
            self.stdout.write(self.style.SUCCESS("All validations passed!"))

            # Ingestion phase
            self.stdout.write("\n=== Ingestion Phase ===")
            created_products = []
            with transaction.atomic():
                created_products = self._ingest_products(
                    products,
                    product_type_map,
                    category_map,
                    attribute_map,
                    channels,
                    warehouse,
                    exchange_rates,
                    moq_value,
                )

            # Update discounted prices after transaction commits
            self.stdout.write("\nUpdating discounted prices...")
            update_discounted_prices_for_promotion(
                ProductModel.objects.filter(id__in=[p.id for p in created_products])
            )
            self.stdout.write(self.style.SUCCESS("✓ Discounted prices updated"))

            self.stdout.write(
                self.style.SUCCESS("\nProduct ingestion completed successfully!")
            )

            self.stdout.write(
                self.style.SUCCESS("\nProduct ingestion completed successfully!")
            )

        except FileNotFoundError as e:
            raise CommandError(f"File not found: {excel_file}") from e
        except Exception as e:
            logger.exception("Error processing Excel file")
            raise CommandError(f"Error processing Excel file: {str(e)}") from e

    def _extract_images(self, excel_file, sheet_name):
        """Extract images from Excel file and map them to row numbers.

        Returns a dict mapping row_number -> image_data.
        """
        workbook = openpyxl.load_workbook(excel_file)
        sheet = workbook[sheet_name]

        images_by_row = {}

        # Get all images from the sheet
        if hasattr(sheet, "_images"):
            for image in sheet._images:
                # Get the image data
                image_data = image._data()

                # Determine which row this image belongs to
                # Images have an anchor attribute that tells us their position
                if hasattr(image, "anchor") and hasattr(image.anchor, "_from"):
                    row_num = image.anchor._from.row + 1  # Excel rows are 1-indexed
                    images_by_row[row_num] = {
                        "data": image_data,
                        "format": image.format,
                    }
                    self.stdout.write(f"Found image for row {row_num}")

        return images_by_row

    def _parse_sizes_and_qty(self, sizes_str, qty_str):
        """Parse sizes and quantities from strings.

        Handles format: "6.5[1], 7[1], 7.5[9], 8[13]"
        where the number in brackets is the quantity for each size.
        """
        if not sizes_str:
            return (), ()

        sizes = []
        quantities = []

        # Split by comma
        parts = sizes_str.split(",")

        for part in parts:
            part = part.strip()
            if not part:
                continue

            # Extract size and quantity using regex: size[qty]
            match = re.match(r"^([^[]+)\[(\d+)\]$", part)
            if match:
                size = match.group(1).strip()
                qty = int(match.group(2))
                sizes.append(size)
                quantities.append(qty)
            else:
                # If no bracket notation, just take the size with 0 quantity
                sizes.append(part)
                quantities.append(0)

        return tuple(sizes), tuple(quantities)

    def _detect_currency(self, row):
        """Detect currency from price fields by looking for currency symbols.

        Returns currency code (GBP, USD, EUR) and logs the decision.
        """
        # Check RRP and Price columns for currency symbols
        rrp_str = str(row["RRP"]) if pd.notna(row["RRP"]) else ""
        price_str = str(row["Price"]) if pd.notna(row["Price"]) else ""

        # Combine both fields to check for symbols
        combined = rrp_str + " " + price_str

        currency = None
        if "£" in combined:
            currency = "GBP"
        elif "$" in combined:
            currency = "USD"
        elif "€" in combined:
            currency = "EUR"
        else:
            currency = "GBP"  # Default fallback
            logger.info(
                "No currency symbol found for product %s, defaulting to GBP",
                row.get("Code", "Unknown"),
            )

        return currency

    def _process_row(self, row, image_data=None):
        """Process a single row from the Excel file and return ProductData instance.

        Expected columns: Image, Code, Description, Category, Sizes, Brand, Qty, RRP, Price
        """
        code = str(row["Code"]).strip() if pd.notna(row["Code"]) else None
        description = (
            str(row["Description"]).strip() if pd.notna(row["Description"]) else ""
        )
        category = str(row["Category"]).strip() if pd.notna(row["Category"]) else ""
        sizes_str = str(row["Sizes"]).strip() if pd.notna(row["Sizes"]) else ""
        qty_str = str(row["Qty"]).strip() if pd.notna(row["Qty"]) else ""
        brand = str(row["Brand"]).strip().title() if pd.notna(row["Brand"]) else ""

        # Extract numeric values from price fields (remove currency symbols)
        rrp_str = str(row["RRP"]) if pd.notna(row["RRP"]) else "0"
        price_str = str(row["Price"]) if pd.notna(row["Price"]) else "0"

        # Remove currency symbols and parse as float
        rrp = (
            float("".join(c for c in rrp_str if c.isdigit() or c == "."))
            if any(c.isdigit() for c in rrp_str)
            else 0.0
        )
        price = (
            float("".join(c for c in price_str if c.isdigit() or c == "."))
            if any(c.isdigit() for c in price_str)
            else 0.0
        )

        if not code:
            self.stdout.write(self.style.WARNING("Skipping row with missing Code"))
            return None

        # Parse sizes and quantities
        sizes, quantities = self._parse_sizes_and_qty(sizes_str, qty_str)

        # Detect currency from price fields
        currency = self._detect_currency(row)
        logger.info("Product %s: Currency detected as %s", code, currency)

        # Create ProductData instance
        product_data = ProductData(
            product_code=code,
            description=description,
            category=category,
            sizes=sizes,
            qty=quantities,
            brand=brand,
            rrp=rrp,
            price=price,
            currency=currency,
            image_data=image_data,
        )

        return product_data

    def _deduplicate_products(self, products):
        """De-duplicate products by product code.

        - If (sizes, qty) tuples are identical: delete duplicates
        - If different: merge sizes/qty, keep highest price and its RRP
        """
        from collections import defaultdict

        # Group by product code
        code_to_products = defaultdict(list)
        for product in products:
            code_to_products[product.product_code].append(product)

        deduplicated = []
        deleted_count = 0
        merged_count = 0

        for code, product_list in code_to_products.items():
            if len(product_list) == 1:
                # No duplicates
                deduplicated.append(product_list[0])
                continue

            # Found duplicates
            self.stdout.write(f"\nFound {len(product_list)} duplicates for code {code}")

            # Check if all have identical (sizes, qty) tuples
            first_tuple = (product_list[0].sizes, product_list[0].qty)
            all_identical = all((p.sizes, p.qty) == first_tuple for p in product_list)

            if all_identical:
                # All identical - keep first, delete rest
                deduplicated.append(product_list[0])
                deleted_count += len(product_list) - 1
                self.stdout.write(
                    f"  → All identical, keeping 1, deleting {len(product_list) - 1}"
                )
            else:
                # Different - merge them
                merged_product = self._merge_products(product_list)
                deduplicated.append(merged_product)
                merged_count += 1
                self.stdout.write(
                    f"  → Merging {len(product_list)} products with different sizes/quantities"
                )

        if deleted_count > 0 or merged_count > 0:
            self.stdout.write("\n=== De-duplication Summary ===")
            if deleted_count > 0:
                self.stdout.write(f"  Deleted {deleted_count} identical duplicate(s)")
            if merged_count > 0:
                self.stdout.write(
                    f"  Merged {merged_count} product(s) with different sizes"
                )

        return deduplicated

    def _merge_products(self, product_list):
        """Merge multiple products with same code but different sizes/quantities.

        Keep highest price and corresponding RRP.
        Combine all sizes and quantities.
        """
        # Find product with highest price
        highest_price_product = max(product_list, key=lambda p: p.price)

        # Combine all sizes and quantities
        all_sizes = []
        all_qty = []
        for product in product_list:
            all_sizes.extend(product.sizes)
            all_qty.extend(product.qty)

        # Create merged product
        merged = ProductData(
            product_code=highest_price_product.product_code,
            description=highest_price_product.description,
            category=highest_price_product.category,
            sizes=tuple(all_sizes),
            qty=tuple(all_qty),
            brand=highest_price_product.brand,
            rrp=highest_price_product.rrp,
            price=highest_price_product.price,
            currency=highest_price_product.currency,
            image_data=highest_price_product.image_data,
        )

        self.stdout.write(
            f"    Keeping price {highest_price_product.price} {highest_price_product.currency}"
        )
        self.stdout.write(f"    Combined to {len(all_sizes)} variants")

        return merged

    def _validate_product_types(self, products):
        """Validate that all categories exist as ProductTypes.

        Returns a mapping of category_name -> ProductType object.
        Raises CommandError if any category is missing.
        """
        # Get unique categories from products
        unique_categories = {p.category for p in products if p.category}

        self.stdout.write(
            f"Validating {len(unique_categories)} unique categories against ProductTypes..."
        )

        # Fetch all matching ProductTypes
        product_types = ProductType.objects.filter(name__in=unique_categories)
        product_type_map = {pt.name: pt for pt in product_types}

        # Check for missing categories
        missing = unique_categories - set(product_type_map.keys())
        if missing:
            raise CommandError(
                f"ProductType validation failed! Missing ProductTypes: {', '.join(sorted(missing))}"
            )

        self.stdout.write(self.style.SUCCESS("✓ All categories found in ProductTypes"))
        return product_type_map

    def _validate_categories(self, products):
        """Validate that all categories exist as Categories.

        Returns a mapping of category_name -> Category object.
        Raises CommandError if any category is missing.
        """
        # Get unique categories from products
        unique_categories = {p.category for p in products if p.category}

        self.stdout.write(
            f"Validating {len(unique_categories)} unique categories against Categories..."
        )

        # Fetch all matching Categories
        categories = Category.objects.filter(name__in=unique_categories)
        category_map = {cat.name: cat for cat in categories}

        # Check for missing categories
        missing = unique_categories - set(category_map.keys())
        if missing:
            raise CommandError(
                f"Category validation failed! Missing Categories: {', '.join(sorted(missing))}"
            )

        self.stdout.write(self.style.SUCCESS("✓ All categories found in Categories"))
        return category_map

    def _validate_attributes(self):
        """Validate that required attributes exist.

        Validates: RRP, Product Code, Size, Minimum Order Quantity, Brand.
        Returns a mapping of attribute_name -> Attribute object.
        Raises CommandError if any attribute is missing.
        """
        required_attributes = [
            "RRP",
            "Product Code",
            "Size",
            "Minimum Order Quantity",
            "Brand",
        ]

        self.stdout.write(
            f"Validating required attributes: {', '.join(required_attributes)}..."
        )

        # Fetch all matching Attributes
        attributes = Attribute.objects.filter(name__in=required_attributes)
        attribute_map = {attr.name: attr for attr in attributes}

        # Check for missing attributes
        missing = set(required_attributes) - set(attribute_map.keys())
        if missing:
            raise CommandError(
                f"Attribute validation failed! Missing Attributes: {', '.join(sorted(missing))}"
            )

        self.stdout.write(self.style.SUCCESS("✓ All required attributes found"))
        return attribute_map

    def _get_minimum_order_quantity(self):
        """Prompt user to enter the minimum order quantity for this import batch.

        Returns the MOQ value as an integer.
        """
        self.stdout.write(self.style.SUCCESS("\n=== Minimum Order Quantity ==="))
        self.stdout.write(
            "Enter the minimum order quantity to apply to all products in this import."
        )

        while True:
            try:
                moq_input = input("Minimum Order Quantity: ").strip()
                moq_value = int(moq_input)
                if moq_value <= 0:
                    self.stdout.write(
                        self.style.ERROR("MOQ must be a positive integer")
                    )
                    continue

                self.stdout.write(self.style.SUCCESS(f"✓ MOQ set to: {moq_value}"))
                return moq_value

            except ValueError:
                self.stdout.write(
                    self.style.ERROR("Invalid input. Please enter a positive integer")
                )
            except KeyboardInterrupt:
                raise CommandError("MOQ input cancelled") from None

    def _validate_product_type_attributes(self, product_type_map, attribute_map):
        """Validate that all required attributes are properly assigned to each product type.

        Product-level: Product Code, RRP, Minimum Order Quantity, Brand
        Variant-level: Size

        Also validates that attributes are NOT in the wrong level.
        """
        self.stdout.write("\nValidating product type attribute assignments...")

        product_level_attrs = ["Product Code", "RRP", "Minimum Order Quantity", "Brand"]
        variant_level_attrs = ["Size"]

        errors = []

        for _category_name, product_type in product_type_map.items():
            self.stdout.write(f"\n  Checking product type: {product_type.name}")

            # Check product-level attributes
            assigned_product_attrs = set(
                product_type.attributeproduct.values_list("attribute__name", flat=True)
            )
            missing_product_attrs = set(product_level_attrs) - assigned_product_attrs

            if missing_product_attrs:
                error_msg = f"Product type '{product_type.name}' is missing product attributes: {', '.join(missing_product_attrs)}"
                errors.append(error_msg)
                self.stdout.write(self.style.ERROR(f"    ✗ {error_msg}"))

            # Check variant-level attributes
            assigned_variant_attrs = set(
                product_type.attributevariant.values_list("attribute__name", flat=True)
            )
            missing_variant_attrs = set(variant_level_attrs) - assigned_variant_attrs

            if missing_variant_attrs:
                error_msg = f"Product type '{product_type.name}' is missing variant attributes: {', '.join(missing_variant_attrs)}"
                errors.append(error_msg)
                self.stdout.write(self.style.ERROR(f"    ✗ {error_msg}"))

            # Check for misplaced attributes - product-level attrs should NOT be in variant-level
            misplaced_in_variants = set(product_level_attrs) & assigned_variant_attrs
            if misplaced_in_variants:
                error_msg = f"Product type '{product_type.name}' has product-level attributes incorrectly assigned as variant attributes: {', '.join(misplaced_in_variants)}"
                errors.append(error_msg)
                self.stdout.write(self.style.ERROR(f"    ✗ {error_msg}"))
                self.stdout.write(
                    self.style.ERROR(
                        "      These should be in Product Attributes, not Variant Attributes!"
                    )
                )

            # Check for misplaced attributes - variant-level attrs should NOT be in product-level
            misplaced_in_products = set(variant_level_attrs) & assigned_product_attrs
            if misplaced_in_products:
                error_msg = f"Product type '{product_type.name}' has variant-level attributes incorrectly assigned as product attributes: {', '.join(misplaced_in_products)}"
                errors.append(error_msg)
                self.stdout.write(self.style.ERROR(f"    ✗ {error_msg}"))
                self.stdout.write(
                    self.style.ERROR(
                        "      These should be in Variant Attributes, not Product Attributes!"
                    )
                )

            if (
                not missing_product_attrs
                and not missing_variant_attrs
                and not misplaced_in_variants
                and not misplaced_in_products
            ):
                self.stdout.write(
                    self.style.SUCCESS("    ✓ All attributes configured correctly")
                )

        if errors:
            self.stdout.write(f"\n{'=' * 60}")
            raise CommandError(
                f"\nProduct type attribute validation failed! {len(errors)} issue(s) found:\n"
                + "\n".join(f"  • {error}" for error in errors)
                + "\n\nPlease configure the attributes correctly in Saleor admin."
            )

        self.stdout.write(
            self.style.SUCCESS(
                "\n✓ All product types have required attributes configured correctly"
            )
        )

    def _validate_sizes(self, products, size_attribute):
        """Validate that all sizes exist as AttributeValues for the Size attribute.

        Prompts user for confirmation if new sizes need to be created.
        """
        # Collect all unique sizes from products
        all_sizes = set()
        for product in products:
            all_sizes.update(product.sizes)

        self.stdout.write(
            f"Found {len(all_sizes)} unique sizes in Excel: {', '.join(sorted(all_sizes))}"
        )

        # Fetch existing size values
        existing_sizes = AttributeValue.objects.filter(attribute=size_attribute)
        existing_size_names = {sv.name for sv in existing_sizes}

        self.stdout.write(
            f"Found {len(existing_size_names)} existing sizes in database"
        )

        # Check for new sizes
        new_sizes = all_sizes - existing_size_names

        if new_sizes:
            self.stdout.write(
                self.style.WARNING(f"\n{len(new_sizes)} new sizes need to be created:")
            )
            for size in sorted(new_sizes):
                self.stdout.write(f"  - {size}")

            # Prompt user for confirmation
            response = input("\nAllow creation of new sizes? [Y/n]: ").strip().lower()
            if response not in ["y", "yes", ""]:
                raise CommandError(
                    "User declined creation of new sizes. Aborting ingestion."
                )

            self.stdout.write(
                self.style.SUCCESS("✓ User approved creation of new sizes")
            )
        else:
            self.stdout.write(self.style.SUCCESS("✓ All sizes already exist"))

    def _validate_unique_product_codes(self, products):
        """Validate that product codes are unique within the Excel file."""
        self.stdout.write("\nValidating product code uniqueness...")

        # Group products by code
        code_to_products: dict[str, list[ProductData]] = {}
        for product in products:
            code = product.product_code
            if code not in code_to_products:
                code_to_products[code] = []
            code_to_products[code].append(product)

        # Check for duplicates
        duplicates = {
            code: prods for code, prods in code_to_products.items() if len(prods) > 1
        }

        if duplicates:
            self.stdout.write(
                self.style.ERROR(
                    f"\n✗ Found {len(duplicates)} duplicate product codes in Excel:"
                )
            )
            for code, prods in duplicates.items():
                self.stdout.write(
                    f"\n  Product Code '{code}' appears {len(prods)} times:"
                )
                for prod in prods:
                    self.stdout.write(f"    - {prod.description}")

            raise CommandError(
                f"\nProduct code validation failed! Found {len(duplicates)} duplicate code(s). "
                f"Each product code must be unique in the Excel file."
            )

        self.stdout.write(
            self.style.SUCCESS(f"✓ All {len(products)} product codes are unique")
        )

    def _validate_unique_names(self, products):
        """Validate that product names (descriptions) are unique.

        Checks uniqueness both within the Excel and against existing products in the database.
        """
        self.stdout.write("\nValidating product name uniqueness...")

        # Generate slugs for all products
        slug_to_products: dict[str, list[ProductData]] = {}
        for product in products:
            slug = slugify(product.description)
            if slug not in slug_to_products:
                slug_to_products[slug] = []
            slug_to_products[slug].append(product)

        # Check for duplicates within Excel
        duplicates_in_excel = {
            slug: prods for slug, prods in slug_to_products.items() if len(prods) > 1
        }

        if duplicates_in_excel:
            self.stdout.write(
                self.style.ERROR(
                    f"\n✗ Found {len(duplicates_in_excel)} duplicate product names in Excel:"
                )
            )
            for slug, prods in duplicates_in_excel.items():
                self.stdout.write(f"\n  Slug '{slug}' would be created from:")
                for prod in prods:
                    self.stdout.write(
                        f"    - Code: {prod.product_code}, Name: {prod.description}"
                    )

        # Check against existing products in database
        all_slugs = list(slug_to_products.keys())
        existing_products = ProductModel.objects.filter(slug__in=all_slugs).values_list(
            "slug", "name"
        )
        existing_slugs = dict(existing_products)

        if existing_slugs:
            self.stdout.write(
                self.style.ERROR(
                    f"\n✗ Found {len(existing_slugs)} product names that already exist in database:"
                )
            )
            for slug, existing_name in existing_slugs.items():
                excel_products = slug_to_products[slug]
                self.stdout.write(f"\n  Slug '{slug}':")
                self.stdout.write(f"    Existing in DB: {existing_name}")
                self.stdout.write("    In Excel:")
                for prod in excel_products:
                    self.stdout.write(
                        f"      - Code: {prod.product_code}, Name: {prod.description}"
                    )

        # Raise error if any duplicates found
        if duplicates_in_excel or existing_slugs:
            total_issues = len(duplicates_in_excel) + len(existing_slugs)
            raise CommandError(
                f"\nProduct name validation failed! Found {total_issues} uniqueness issue(s). "
                f"Please fix duplicate names in the Excel file or remove products that already exist in the database."
            )

        self.stdout.write(
            self.style.SUCCESS(f"✓ All {len(products)} product names are unique")
        )

    def _confirm_price_interpretation(self):
        """Confirm with user that the Price column is the sale price and does NOT include VAT."""
        self.stdout.write(
            self.style.WARNING("\n=== Price Interpretation Confirmation ===")
        )
        self.stdout.write("IMPORTANT: Please confirm the following:")
        self.stdout.write(
            "  1. The 'Price' column contains the SALE PRICE (the price customers pay)"
        )
        self.stdout.write("  2. The 'Price' column does NOT include VAT")
        self.stdout.write("")

        response = input("Is this correct? [Y/n]: ").strip().lower()
        if response not in ["y", "yes", ""]:
            raise CommandError(
                "User did not confirm price interpretation. Aborting ingestion."
            )

        self.stdout.write(self.style.SUCCESS("✓ Price interpretation confirmed"))

    def _validate_and_fetch_channels(self):
        """Fetch all active channels and their currencies.

        Returns a list of Channel objects.
        """
        channels = Channel.objects.filter(is_active=True)

        if not channels.exists():
            raise CommandError("No active channels found in the database!")

        self.stdout.write(self.style.SUCCESS("\n=== Active Channels ==="))
        self.stdout.write(f"Found {len(channels)} active channel(s):")
        for channel in channels:
            self.stdout.write(f"  • {channel.slug} ({channel.name})")
            self.stdout.write(f"    Currency: {channel.currency_code}")
            self.stdout.write(f"    Country: {channel.default_country}")

        return list(channels)

    def _select_warehouse(self):
        """Select warehouse for stock allocation.

        If multiple warehouses exist, prompt user to select one.
        If only one exists, use it automatically.
        """
        warehouses = Warehouse.objects.all()

        if not warehouses.exists():
            raise CommandError("No warehouses found in the database!")

        if len(warehouses) == 1:
            warehouse = warehouses.first()
            if warehouse is None:
                raise CommandError("No warehouse found")
            self.stdout.write(self.style.SUCCESS("\n=== Warehouse ==="))
            self.stdout.write(f"Using warehouse: {warehouse.slug} ({warehouse.name})")
            return warehouse

        # Multiple warehouses - prompt user to select
        self.stdout.write(self.style.SUCCESS("\n=== Warehouse Selection ==="))
        self.stdout.write(f"Found {len(warehouses)} warehouses:")
        warehouse_list = list(warehouses)
        for idx, warehouse in enumerate(warehouse_list, 1):
            self.stdout.write(f"  {idx}. {warehouse.slug} ({warehouse.name})")

        while True:
            try:
                choice = input(
                    f"\nSelect warehouse (1-{len(warehouse_list)}): "
                ).strip()
                warehouse_idx = int(choice) - 1
                if 0 <= warehouse_idx < len(warehouse_list):
                    selected_warehouse = warehouse_list[warehouse_idx]
                    self.stdout.write(
                        self.style.SUCCESS(
                            f"✓ Selected warehouse: {selected_warehouse.slug}"
                        )
                    )
                    return selected_warehouse
                self.stdout.write(
                    self.style.ERROR(
                        f"Invalid choice. Please enter 1-{len(warehouse_list)}"
                    )
                )
            except (ValueError, KeyboardInterrupt):
                raise CommandError("Warehouse selection cancelled") from None

    def _fetch_exchange_rates(self, products, channels):
        """Fetch current exchange rates for all currencies needed.

        Uses frankfurter.app API (European Central Bank rates, free).
        Returns a dict: {from_currency: {to_currency: rate}}
        """
        # Get unique currencies from products
        product_currencies = {p.currency for p in products}

        # Get currencies from channels
        channel_currencies = {ch.currency_code for ch in channels}

        self.stdout.write(self.style.SUCCESS("\n=== Exchange Rates ==="))
        self.stdout.write(
            f"Product currencies in Excel: {', '.join(sorted(product_currencies))}"
        )
        self.stdout.write(
            f"Channel currencies needed: {', '.join(sorted(channel_currencies))}"
        )
        self.stdout.write("\nFetching current rates from European Central Bank...")

        exchange_rates: dict[str, dict[str, Decimal]] = {}

        # Fetch rates for each product currency to all channel currencies
        for from_currency in product_currencies:
            exchange_rates[from_currency] = {}

            # Add 1:1 rate for same currency
            exchange_rates[from_currency][from_currency] = Decimal("1.0")

            # Fetch rates for other currencies
            other_currencies = channel_currencies - {from_currency}
            if other_currencies:
                try:
                    # Frankfurter API: https://www.frankfurter.app/
                    url = f"https://api.frankfurter.app/latest?from={from_currency}&to={','.join(other_currencies)}"
                    response = HTTPClient.send_request(
                        "GET", url, timeout=10, allow_redirects=False
                    )
                    response.raise_for_status()
                    data = response.json()

                    for to_currency, rate in data["rates"].items():
                        exchange_rates[from_currency][to_currency] = Decimal(str(rate))

                except Exception as e:
                    logger.exception("Failed to fetch exchange rates")
                    raise CommandError(
                        f"Failed to fetch exchange rates: {str(e)}"
                    ) from e

        # Log all rates that will be used
        self.stdout.write("\nRates that will be used for price conversion:")
        for from_curr in sorted(exchange_rates.keys()):
            for to_curr in sorted(exchange_rates[from_curr].keys()):
                rate = exchange_rates[from_curr][to_curr]
                if from_curr == to_curr:
                    self.stdout.write(
                        f"  • {from_curr} → {to_curr}: {rate} (same currency)"
                    )
                else:
                    self.stdout.write(f"  • {from_curr} → {to_curr}: {rate}")
                    # Show example conversion
                    example_amount = Decimal("100.00")
                    converted = example_amount * rate
                    self.stdout.write(
                        f"    Example: {from_curr} {example_amount} = {to_curr} {converted:.2f}"
                    )

        self.stdout.write(self.style.SUCCESS("\n✓ Exchange rates fetched successfully"))
        return exchange_rates

    def _convert_price(self, amount, from_currency, to_currency, exchange_rates):
        """Convert price from one currency to another."""
        if from_currency == to_currency:
            return Decimal(str(amount))

        rate = exchange_rates.get(from_currency, {}).get(to_currency)
        if rate is None:
            raise CommandError(
                f"No exchange rate found for {from_currency} → {to_currency}"
            )

        return Decimal(str(amount)) * rate

    def _ingest_products(
        self,
        products,
        product_type_map,
        category_map,
        attribute_map,
        channels,
        warehouse,
        exchange_rates,
        moq_value,
    ):
        """Ingest products into the database within a transaction.

        Creates products with variants for each size, and prices for each channel currency.
        """
        self.stdout.write(f"\nIngesting {len(products)} products...")
        self.stdout.write(
            self.style.WARNING("All products will be created in a single transaction.")
        )
        self.stdout.write(
            self.style.WARNING(
                "If ANY product fails, ALL changes will be rolled back.\n"
            )
        )

        created_count = 0
        created_products = []

        for idx, product_data in enumerate(products, 1):
            self.stdout.write(
                f"\n[{idx}/{len(products)}] Processing: {product_data.product_code} - {product_data.description}"
            )

            # 1. Create Product
            product = self._create_product(product_data, product_type_map, category_map)

            # 2. Create ProductChannelListings for all channels
            for channel in channels:
                self._create_product_channel_listing(product, channel)

            # 3. Create ProductMedia (if image exists)
            if product_data.image_data:
                self._create_product_media(product, product_data)

            # 4. Assign product-level attributes (Product Code, RRP, MOQ, Brand)
            self._assign_product_attributes(
                product, product_data, attribute_map, moq_value
            )

            # 5. Create variants for each size
            variant_count = 0
            for size, qty in zip(product_data.sizes, product_data.qty, strict=False):
                variant = self._create_variant(product, size)

                # 6. Assign variant-level attributes (Size)
                self._assign_variant_attributes(
                    variant,
                    size,
                    attribute_map,
                    product_type_map[product_data.category],
                )

                # 7. Create channel listings with converted prices
                for channel in channels:
                    self._create_variant_channel_listing(
                        variant, channel, product_data, exchange_rates
                    )

                # 8. Create stock
                self._create_stock(variant, warehouse, qty)
                variant_count += 1

            self.stdout.write(
                self.style.SUCCESS(
                    f"  ✓ Created product: {product.name} ({variant_count} variants)"
                )
            )
            created_products.append(product)
            created_count += 1

        # All good - transaction will commit
        self.stdout.write(f"\n{'=' * 60}")
        self.stdout.write(
            self.style.SUCCESS(f"✓ Successfully created {created_count} products!")
        )

        return created_products

    def _create_product(self, product_data, product_type_map, category_map):
        """Create a Product instance."""
        # Generate slug from description (product name)
        slug = slugify(product_data.description)

        product = ProductModel.objects.create(
            name=product_data.description,  # Use Description column as product name
            slug=slug,
            product_type=product_type_map[product_data.category],
            category=category_map[product_data.category],
        )
        return product

    def _create_product_channel_listing(self, product, channel):
        """Create ProductChannelListing to make product available in a channel."""
        from django.utils import timezone

        ProductChannelListing.objects.create(
            product=product,
            channel=channel,
            currency=channel.currency_code,
            is_published=True,
            visible_in_listings=True,
            available_for_purchase_at=timezone.now(),
        )

    def _create_product_media(self, product, product_data):
        """Create ProductMedia with the image."""
        image_data = product_data.image_data["data"]
        image_format = product_data.image_data["format"]

        # Create ContentFile from image bytes
        image_file = ContentFile(image_data)
        filename = f"{product_data.product_code}.{image_format}"

        media = ProductMedia.objects.create(
            product=product,
            alt=product_data.product_code,
        )
        media.image.save(filename, image_file, save=True)

    def _assign_product_attributes(
        self, product, product_data, attribute_map, moq_value
    ):
        """Assign Product Code, RRP, MOQ, and Brand attributes to the product."""
        # Product Code attribute
        product_code_attr = attribute_map["Product Code"]
        product_code_slug = slugify(product_data.product_code)
        try:
            product_code_value = AttributeValue.objects.get(
                attribute=product_code_attr, slug=product_code_slug
            )
        except AttributeValue.DoesNotExist:
            product_code_value, _ = AttributeValue.objects.get_or_create(
                attribute=product_code_attr,
                name=product_data.product_code,
                defaults={"slug": product_code_slug},
            )

        AssignedProductAttributeValue.objects.create(
            product=product, value=product_code_value
        )

        # RRP attribute
        rrp_attr = attribute_map["RRP"]
        rrp_slug = slugify(f"rrp-{product_data.rrp}")
        try:
            rrp_value = AttributeValue.objects.get(attribute=rrp_attr, slug=rrp_slug)
        except AttributeValue.DoesNotExist:
            rrp_value, _ = AttributeValue.objects.get_or_create(
                attribute=rrp_attr,
                name=str(product_data.rrp),
                defaults={"slug": rrp_slug},
            )

        AssignedProductAttributeValue.objects.create(product=product, value=rrp_value)

        # Minimum Order Quantity attribute
        moq_attr = attribute_map["Minimum Order Quantity"]
        moq_slug = slugify(f"moq-{moq_value}")
        try:
            moq_value_obj = AttributeValue.objects.get(
                attribute=moq_attr, slug=moq_slug
            )
        except AttributeValue.DoesNotExist:
            moq_value_obj, _ = AttributeValue.objects.get_or_create(
                attribute=moq_attr, name=str(moq_value), defaults={"slug": moq_slug}
            )

        AssignedProductAttributeValue.objects.create(
            product=product, value=moq_value_obj
        )

        # Brand attribute
        brand_attr = attribute_map["Brand"]
        brand_slug = slugify(product_data.brand)
        try:
            brand_value = AttributeValue.objects.get(
                attribute=brand_attr, slug=brand_slug
            )
        except AttributeValue.DoesNotExist:
            brand_value, _ = AttributeValue.objects.get_or_create(
                attribute=brand_attr,
                name=product_data.brand,
                defaults={"slug": brand_slug},
            )

        AssignedProductAttributeValue.objects.create(product=product, value=brand_value)

    def _create_variant(self, product, size):
        """Create a ProductVariant for a specific size."""
        variant = ProductVariant.objects.create(
            product=product,
            name=size,
            sku=None,  # Leave SKU null for auto-population
        )
        return variant

    def _assign_variant_attributes(self, variant, size, attribute_map, product_type):
        """Assign Size attribute to the variant."""
        size_attr = attribute_map["Size"]

        # Get or create the AttributeVariant (the link between attribute and product type)
        attr_variant, _ = AttributeVariant.objects.get_or_create(
            attribute=size_attr, product_type=product_type
        )

        # Get or create the size value - try to find existing first by slug
        size_slug = slugify(f"size-{size}")
        try:
            # Try to find by slug and attribute first (most likely to exist)
            size_value = AttributeValue.objects.get(attribute=size_attr, slug=size_slug)
        except AttributeValue.DoesNotExist:
            # Doesn't exist, create it
            size_value, _ = AttributeValue.objects.get_or_create(
                attribute=size_attr, name=size, defaults={"slug": size_slug}
            )

        # Create the assignment
        assigned_attr, _ = AssignedVariantAttribute.objects.get_or_create(
            variant=variant, assignment=attr_variant
        )

        # Link the value
        AssignedVariantAttributeValue.objects.create(
            value=size_value, assignment=assigned_attr, variant=variant
        )

    def _create_variant_channel_listing(
        self, variant, channel, product_data, exchange_rates
    ):
        """Create ProductVariantChannelListing with converted price for the channel."""
        # Convert price to channel currency
        converted_price = self._convert_price(
            product_data.price,
            product_data.currency,
            channel.currency_code,
            exchange_rates,
        )

        ProductVariantChannelListing.objects.create(
            variant=variant,
            channel=channel,
            currency=channel.currency_code,
            price_amount=converted_price,
        )

    def _create_stock(self, variant, warehouse, quantity):
        """Create Stock entry for the variant in the warehouse."""
        Stock.objects.create(
            warehouse=warehouse,
            product_variant=variant,
            quantity=quantity,
        )
